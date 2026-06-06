from __future__ import annotations
import io
import asyncio
import aiosqlite
from aiogram import Router, Bot
from aiogram.filters import Command
from aiogram.types import Message, BufferedInputFile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.config.settings import settings

router = Router()

@router.message(Command("admin_export"))
async def admin_export_stats(m: Message, bot: Bot, db: aiosqlite.Connection):
    """
    Export global user stats and account details into a multi-sheet Excel file.
    Accessible only to Telegram IDs present in ADMIN_IDS.
    """
    # 1. Security Check
    if m.from_user.id not in settings.admin_ids:
        # Ignore silently for security
        return

    # 2. SQL Queries (validated and correct fields)
    users_query = """
        SELECT 
            u.id, 
            u.full_access, 
            u.created_at,
            s.lang, 
            s.timezone,
            (SELECT COUNT(*) FROM accounts a WHERE a.user_id = u.id AND a.is_archived = 0) as accounts_count,
            (SELECT COUNT(*) FROM transactions t WHERE t.user_id = u.id AND t.deleted_at IS NULL) as tx_count,
            (SELECT COUNT(*) FROM debts d WHERE d.user_id = u.id AND d.closed_at IS NULL) as active_debts,
            u.telegram_id,
            u.username,
            u.display_name
        FROM users u
        LEFT JOIN settings s ON u.id = s.user_id
        ORDER BY u.created_at DESC
    """
    
    accounts_query = """
        SELECT 
            a.user_id, 
            a.name, 
            a.balance, 
            a.currency, 
            a.is_saving
        FROM accounts a
        WHERE a.is_archived = 0
        ORDER BY a.user_id, a.name
    """

    try:
        async with db.execute(users_query) as cursor:
            users = await cursor.fetchall()
            
        async with db.execute(accounts_query) as cursor:
            accounts = await cursor.fetchall()
    except Exception as e:
        await m.reply(f"❌ Ошибка выполнения запросов к БД: {e}")
        return

    # 3. Resolve Usernames and Names dynamically from Telegram Bot API
    user_usernames = {}
    user_first_names = {}
    user_tg_ids = {}
    
    for u in users:
        db_id = u[0]
        tg_id = u[8]
        username = u[9]
        display_name = u[10]
        
        user_tg_ids[db_id] = tg_id
        
        if tg_id:
            try:
                chat = await bot.get_chat(tg_id)
                username_val = f"@{chat.username}" if chat.username else (f"@{username}" if username else "—")
                first_name_val = chat.first_name or display_name or "—"
            except Exception:
                username_val = f"@{username}" if username else "—"
                first_name_val = display_name or "—"
        else:
            username_val = f"@{username}" if username else "—"
            first_name_val = display_name or "—"
            
        user_usernames[db_id] = username_val
        user_first_names[db_id] = first_name_val
        if tg_id:
            # Prevent Telegram API spam rate-limits
            await asyncio.sleep(0.05)

    # 4. Create Workbook
    wb = Workbook()
    
    # Fonts and Fills (Premium Deep Blue/Slate theme)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    # ==========================================
    # Sheet 1: Users
    # ==========================================
    ws_users = wb.active
    ws_users.title = "Пользователи"
    
    headers_users = [
        "Telegram ID", "Юзернейм", "Имя", "Premium (Full Access)", 
        "Дата регистрации", "Язык", "Часовой пояс", "Кол-во счетов", 
        "Кол-во операций", "Активные долги"
    ]
    
    ws_users.append(headers_users)
    for col_idx in range(1, len(headers_users) + 1):
        cell = ws_users.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    for row_idx, u in enumerate(users, start=2):
        db_id = u[0]
        tg_id = u[8]
        is_premium = "Да 👑" if u[1] == 1 else "Нет"
        username = user_usernames.get(db_id, "—")
        first_name = user_first_names.get(db_id, "—")
        tg_id_display = tg_id if tg_id else f"App ID: {db_id}"
        
        row_data = [
            tg_id_display, username, first_name, is_premium,
            u[2] or "—", (u[3] or "ru").upper(), u[4] or "—", u[5], u[6], u[7]
        ]
        ws_users.append(row_data)
        for col_idx in range(1, len(headers_users) + 1):
            cell = ws_users.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            # Align numeric fields to right, text to left
            if col_idx in (1, 8, 9, 10):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in (4, 6, 7):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # ==========================================
    # Sheet 2: Accounts
    # ==========================================
    ws_accounts = wb.create_sheet(title="Счета пользователей")
    
    headers_accounts = [
        "Telegram ID владельца", "Юзернейм владельца", "Название счета", 
        "Баланс", "Валюта", "Накопительный?"
    ]
    
    ws_accounts.append(headers_accounts)
    for col_idx in range(1, len(headers_accounts) + 1):
        cell = ws_accounts.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        
    for row_idx, acc in enumerate(accounts, start=2):
        db_id = acc[0]
        is_saving = "Да 🎯" if acc[4] == 1 else "Нет 💳"
        username = user_usernames.get(db_id, "—")
        tg_id = user_tg_ids.get(db_id)
        tg_id_display = tg_id if tg_id else f"App ID: {db_id}"
        
        row_data = [
            tg_id_display, username, acc[1],
            acc[2], acc[3], is_saving
        ]
        ws_accounts.append(row_data)
        for col_idx in range(1, len(headers_accounts) + 1):
            cell = ws_accounts.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = thin_border
            # Align fields
            if col_idx in (1, 4):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col_idx in (5, 6):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-adjust column widths for both sheets
    for ws in (ws_users, ws_accounts):
        ws.row_dimensions[1].height = 28
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or "")
                # Exclude long header text lengths if they blow up width too much
                if cell.row == 1:
                    max_len = max(max_len, min(len(val_str), 15))
                else:
                    max_len = max(max_len, len(val_str))
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
            # Apply padding height to data rows
            for cell in col:
                if cell.row > 1:
                    ws.row_dimensions[cell.row].height = 20

    # 5. Save to buffer and send
    try:
        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        xlsx_bytes = file_stream.read()
    except Exception as e:
        await m.reply(f"❌ Ошибка формирования Excel файла: {e}")
        return

    await m.reply_document(
        BufferedInputFile(xlsx_bytes, filename="admin_global_stats.xlsx"),
        caption="📊 <b>Глобальная статистика пользователей и счетов</b>"
    )


@router.message(Command("admin_user"))
async def admin_user_info(m: Message, bot: Bot, db: aiosqlite.Connection):
    """
    Get detailed diagnostic information for a specific user ID.
    """
    if m.from_user.id not in settings.admin_ids:
        return
        
    parts = m.text.split()
    if len(parts) < 2 or not parts[1].isdigit():
        await m.reply("❌ Использование: <code>/admin_user &lt;user_id&gt;</code>", parse_mode="HTML")
        return
        
    target_id = int(parts[1])
    
    cur = await db.execute("SELECT id, telegram_id, username, display_name, created_at FROM users WHERE id = ? OR telegram_id = ?", (target_id, target_id))
    row = await cur.fetchone()
    if not row:
        await m.reply(f"❌ Пользователь <code>{target_id}</code> не найден в базе данных.", parse_mode="HTML")
        return
        
    db_user_id, telegram_id, stored_username, display_name, created_at = row[0], row[1], row[2], row[3], row[4]
    
    from app.domain.services.access_service import get_user_context, get_available_features
    from app.db.repositories.settings_repo import get_lang
    
    ctx = await get_user_context(db, db_user_id)
    lang = await get_lang(db, db_user_id)
    features = await get_available_features(db, db_user_id)
    
    name = display_name or "—"
    username = f"@{stored_username}" if stored_username else "—"
    
    if telegram_id:
        try:
            chat = await bot.get_chat(telegram_id)
            name = f"{chat.first_name or ''} {chat.last_name or ''}".strip() or name
            username = f"@{chat.username}" if chat.username else username
        except Exception:
            pass
        
    days_left = "—"
    if ctx.expiration_date:
        from datetime import date as _date
        try:
            exp = _date.fromisoformat(ctx.expiration_date)
            from app.domain.time_utils import today_in_user_tz
            today = await today_in_user_tz(db, db_user_id)
            days = (exp - today).days
            days_left = f"{days} дней" if days >= 0 else "истёк"
        except Exception:
            pass
            
    info = (
        f"👤 <b>Информация о пользователе {db_user_id}</b>\n\n"
        f"• <b>Telegram ID:</b> <code>{telegram_id or '—'}</code>\n"
        f"• <b>Имя:</b> {name}\n"
        f"• <b>Юзернейм:</b> {username}\n"
        f"• <b>Регистрация:</b> <code>{created_at}</code>\n"
        f"• <b>Язык:</b> <code>{lang.upper()}</code>\n"
        f"• <b>Режим (mode):</b> <code>{ctx.mode}</code>\n"
        f"• <b>Full Access:</b> <code>{ctx.full_access}</code>\n"
        f"• <b>Истекает:</b> <code>{ctx.expiration_date or '—'}</code> ({days_left})\n"
        f"• <b>Уровень прогресса:</b> <code>{ctx.progress_level}</code>\n"
        f"• <b>Серия:</b> <code>{ctx.current_streak}</code> (макс. <code>{ctx.max_streak}</code>)\n"
        f"• <b>Доступные функции:</b>\n"
        f"<code>{', '.join(sorted(features))}</code>"
    )
    await m.reply(info, parse_mode="HTML")


@router.message(Command("admin_grant"))
async def admin_grant_access(m: Message, bot: Bot, db: aiosqlite.Connection):
    """
    Grant premium (full access) to a user for a specific number of days.
    """
    if m.from_user.id not in settings.admin_ids:
        return
        
    parts = m.text.split()
    if len(parts) < 2 or not parts[1].isdigit():
        await m.reply("❌ Использование: <code>/admin_grant &lt;user_id&gt; [кол-во_дней]</code>", parse_mode="HTML")
        return
        
    target_id = int(parts[1])
    days = 365
    if len(parts) >= 3 and parts[2].isdigit():
        days = int(parts[2])
        
    cur = await db.execute("SELECT id, telegram_id FROM users WHERE id = ? OR telegram_id = ?", (target_id, target_id))
    row = await cur.fetchone()
    if not row:
        await m.reply(f"❌ Пользователь <code>{target_id}</code> не найден в базе данных.", parse_mode="HTML")
        return
        
    db_user_id, telegram_id = row[0], row[1]
        
    from app.db.repositories.users_repo import grant_full_access
    from app.db.repositories.settings_repo import get_lang
    from app.handlers.common import build_main_menu_markup
    from datetime import datetime as _dt, timezone as _tz
    
    try:
        await grant_full_access(db, db_user_id, days=days)
        now_str = _dt.now(_tz.utc).isoformat()
        await db.execute(
            "UPDATE settings SET trial_reminder_sent = 0, updated_at = ? WHERE user_id = ?",
            (now_str, db_user_id)
        )
        await db.commit()
    except Exception as e:
        await db.rollback()
        await m.reply(f"❌ Ошибка записи в базу данных: {e}")
        return
        
    lang = await get_lang(db, db_user_id)
    
    await m.reply(f"✅ Успешно выдан Полный доступ (Premium) пользователю <code>{db_user_id}</code> на <b>{days}</b> дней.", parse_mode="HTML")
    
    if telegram_id:
        user_msg = {
            "ru": f"🎉 <b>Администратор активировал вам Полный режим на {days} дней!</b>\n\nТеперь вам доступны абсолютно все функции бота, включая AI-Консультанта, отчеты по категориям, учет долгов, лимиты и цели.",
            "en": f"🎉 <b>The administrator has activated Full Mode for you for {days} days!</b>\n\nAll features of the bot are now available to you, including the AI Assistant, category reports, debts tracking, limits, and targets.",
            "kk": f"🎉 <b>Әкімші сізге {days} күнге Толық режимді қосты!</b>\n\nЕнді сізге боттың барлық функциялары, соның ішінде AI-Кеңесші, санаттар бойынша есептер, қарыздарды есепке алу, лимиттер мен мақсаттар қолжетімді.",
        }.get(lang, "ru")
        
        try:
            markup = await build_main_menu_markup(db, db_user_id, lang)
            await bot.send_message(telegram_id, user_msg, reply_markup=markup, parse_mode="HTML")
        except Exception as e:
            await m.reply(f"⚠️ Доступ выдан, но не удалось отправить сообщение пользователю: {e}")


@router.message(Command("admin_revoke"))
async def admin_revoke_access(m: Message, bot: Bot, db: aiosqlite.Connection):
    """
    Revoke premium (full access) from a user immediately.
    """
    if m.from_user.id not in settings.admin_ids:
        return
        
    parts = m.text.split()
    if len(parts) < 2 or not parts[1].isdigit():
        await m.reply("❌ Использование: <code>/admin_revoke &lt;user_id&gt;</code>", parse_mode="HTML")
        return
        
    target_id = int(parts[1])
    
    cur = await db.execute("SELECT id, telegram_id FROM users WHERE id = ? OR telegram_id = ?", (target_id, target_id))
    row = await cur.fetchone()
    if not row:
        await m.reply(f"❌ Пользователь <code>{target_id}</code> не найден в базе данных.", parse_mode="HTML")
        return
        
    db_user_id, telegram_id = row[0], row[1]
        
    from app.db.repositories.settings_repo import get_lang
    from app.handlers.common import build_main_menu_markup
    
    try:
        await db.execute(
            "UPDATE users SET full_access = 0, mode = 'newbie' WHERE id = ?",
            (db_user_id,)
        )
        await db.commit()
    except Exception as e:
        await db.rollback()
        await m.reply(f"❌ Ошибка записи в базу данных: {e}")
        return
        
    lang = await get_lang(db, db_user_id)
    
    await m.reply(f"✅ Доступ (Premium) для пользователя <code>{db_user_id}</code> успешно аннулирован.", parse_mode="HTML")
    
    if telegram_id:
        user_msg = {
            "ru": "ℹ️ <b>Ваш Полный доступ был приостановлен или изменен администратором.</b>\n\nБот переведен в стандартный режим.",
            "en": "ℹ️ <b>Your Full access has been suspended or modified by the administrator.</b>\n\nThe bot has been switched to standard mode.",
            "kk": "ℹ️ <b>Сіздің Толық қолжетімділігіңіз әкімшімен тоқтатылды немесе өзгертілді.</b>\n\nБот стандартты режимге ауыстырылды.",
        }.get(lang, "ru")
        
        try:
            markup = await build_main_menu_markup(db, db_user_id, lang)
            await bot.send_message(telegram_id, user_msg, reply_markup=markup, parse_mode="HTML")
        except Exception as e:
            await m.reply(f"⚠️ Доступ аннулирован, но не удалось отправить сообщение пользователю: {e}")


@router.message(Command("info"))
async def admin_info_help(m: Message):
    """
    Cheat sheet for administrative commands.
    """
    if m.from_user.id not in settings.admin_ids:
        return
        
    info_text = (
        "🛠 <b>Панель администратора — Список команд</b>\n\n"
        "Вы можете использовать следующие команды для управления пользователями:\n\n"
        "📊 <b>Статистика и экспорт</b>\n"
        "• <code>/admin_export</code> — Экспортировать глобальную статистику пользователей и их счетов в Excel-файл.\n"
        "• <code>/admin_user &lt;user_id&gt;</code> — Получить подробную диагностическую информацию о пользователе.\n\n"
        "👑 <b>Управление Premium (Полным доступом)</b>\n"
        "• <code>/admin_grant &lt;user_id&gt; [кол-во_дней]</code> — Выдать пользователю Premium (Полный доступ) на указанное количество дней (по умолчанию 365).\n"
        "• <code>/admin_revoke &lt;user_id&gt;</code> — Забрать Premium (Полный доступ) у пользователя и вернуть стандартный режим.\n\n"
        "🔥 <b>Управление серией активности (Streak)</b>\n"
        "• <code>/admin_streak &lt;user_id&gt; &lt;значение&gt;</code> — Установить текущую серию активности пользователя (в днях). Максимальная серия обновится автоматически, если новое значение больше старого.\n\n"
        "🎁 <b>Начисление бесплатных AI-отчетов</b>\n"
        "• <code>/admin_reports &lt;user_id&gt; &lt;количество&gt;</code> — Начислить пользователю дополнительные бесплатные AI-отчеты."
    )
    await m.reply(info_text, parse_mode="HTML")


@router.message(Command("admin_streak"))
async def admin_set_streak(m: Message, bot: Bot, db: aiosqlite.Connection):
    """
    Set user activity streak.
    Usage: /admin_streak <user_id> <streak_value>
    """
    if m.from_user.id not in settings.admin_ids:
        return

    parts = m.text.split()
    if len(parts) < 3 or not parts[1].isdigit() or not parts[2].isdigit():
        await m.reply("❌ Использование: <code>/admin_streak &lt;user_id&gt; &lt;значение&gt;</code>", parse_mode="HTML")
        return

    target_id = int(parts[1])
    streak_value = int(parts[2])

    cur = await db.execute("SELECT id, telegram_id, max_streak FROM users WHERE id = ? OR telegram_id = ?", (target_id, target_id))
    row = await cur.fetchone()
    if not row:
        await m.reply(f"❌ Пользователь <code>{target_id}</code> не найден в базе данных.", parse_mode="HTML")
        return

    db_user_id, telegram_id, max_streak = row[0], row[1], int(row[2] or 0)
    new_max = max(max_streak, streak_value)

    from app.domain.time_utils import today_in_user_tz
    try:
        today = await today_in_user_tz(db, db_user_id)
        today_str = today.isoformat()
    except Exception:
        from datetime import datetime as _dt, timezone as _tz
        today_str = _dt.now(_tz.utc).date().isoformat()

    try:
        await db.execute(
            "UPDATE users SET current_streak = ?, max_streak = ?, last_activity_date = ? WHERE id = ?",
            (streak_value, new_max, today_str, db_user_id)
        )
        await db.commit()
    except Exception as e:
        await db.rollback()
        await m.reply(f"❌ Ошибка записи в базу данных: {e}")
        return

    await m.reply(
        f"✅ Серия активности пользователя <code>{db_user_id}</code> успешно изменена на <b>{streak_value}</b> дней (макс. серия: <b>{new_max}</b>).",
        parse_mode="HTML"
    )

    if telegram_id:
        from app.db.repositories.settings_repo import get_lang
        lang = await get_lang(db, db_user_id)

        user_msg = {
            "ru": f"🔥 <b>Администратор установил вашу серию активности: {streak_value} дн. подряд!</b>\n\nПродолжайте вести учёт каждый день, чтобы сохранить её!",
            "en": f"🔥 <b>The administrator has set your activity streak to {streak_value} days!</b>\n\nKeep tracking your transactions daily to maintain it!",
            "kk": f"🔥 <b>Әкімші сіздің белсенділік серияңызды {streak_value} күн етіп орнатты!</b>\n\nОны сақтап қалу үшін күнделікті шығындарды жазып тұрыңыз!",
        }.get(lang, "ru")

        try:
            await bot.send_message(telegram_id, user_msg, parse_mode="HTML")
        except Exception as e:
            await m.reply(f"⚠️ Серия изменена, но не удалось отправить сообщение пользователю: {e}")


@router.message(Command("admin_reports"))
async def admin_grant_reports(m: Message, bot: Bot, db: aiosqlite.Connection):
    """
    Grant extra free AI reports/credits to a user.
    Usage: /admin_reports <user_id> <count>
    """
    if m.from_user.id not in settings.admin_ids:
        return

    parts = m.text.split()
    if len(parts) < 3 or not parts[1].isdigit() or not parts[2].isdigit():
        await m.reply("❌ Использование: <code>/admin_reports &lt;user_id&gt; &lt;количество&gt;</code>", parse_mode="HTML")
        return

    target_id = int(parts[1])
    count = int(parts[2])

    cur = await db.execute("SELECT id, telegram_id FROM users WHERE id = ? OR telegram_id = ?", (target_id, target_id))
    row = await cur.fetchone()
    if not row:
        await m.reply(f"❌ Пользователь <code>{target_id}</code> не найден в базе данных.", parse_mode="HTML")
        return

    db_user_id, telegram_id = row[0], row[1]

    from app.db.repositories.settings_repo import add_ai_reports_extra, get_lang
    from datetime import datetime as _dt, timezone as _tz
    now_str = _dt.now(_tz.utc).isoformat()

    try:
        await add_ai_reports_extra(db, db_user_id, count, now_str)
        await db.commit()
    except Exception as e:
        await db.rollback()
        await m.reply(f"❌ Ошибка записи в базу данных: {e}")
        return

    await m.reply(
        f"✅ Пользователю <code>{db_user_id}</code> успешно начислено <b>{count}</b> бесплатных AI-отчетов.",
        parse_mode="HTML"
    )

    if telegram_id:
        lang = await get_lang(db, db_user_id)

        user_msg = {
            "ru": f"🎁 <b>Администратор начислил вам бесплатные AI-отчеты: +{count} шт.!</b>\n\nВы можете использовать их в разделе AI-Консультанта.",
            "en": f"🎁 <b>The administrator has credited you with free AI reports: +{count}!</b>\n\nYou can use them in the AI Assistant section.",
            "kk": f"🎁 <b>Әкімші сізге тегін AI-есептерді қосты: +{count} дана!</b>\n\nОларды AI-Кеңесші бөлімінде пайдалана аласыз.",
        }.get(lang, "ru")

        try:
            await bot.send_message(telegram_id, user_msg, parse_mode="HTML")
        except Exception as e:
            await m.reply(f"⚠️ Отчеты начислены, но не удалось отправить сообщение пользователю: {e}")


@router.message(Command("admin_delete"))
async def admin_delete_user(m: Message, db: aiosqlite.Connection):
    """
    Completely deletes a user and all their records from all database tables.
    Usage: /admin_delete <user_id>
    """
    if m.from_user.id not in settings.admin_ids:
        return

    parts = m.text.split()
    if len(parts) < 2 or not parts[1].isdigit():
        await m.reply("❌ Использование: <code>/admin_delete &lt;user_id&gt;</code>", parse_mode="HTML")
        return

    target_id = int(parts[1])

    # 1. Verify user exists
    cur = await db.execute("SELECT id FROM users WHERE id = ? OR telegram_id = ?", (target_id, target_id))
    row = await cur.fetchone()
    if not row:
        await m.reply(f"❌ Пользователь <code>{target_id}</code> не найден в базе данных.", parse_mode="HTML")
        return

    db_user_id = row[0]

    try:
        from app.db.repositories.reset_repo import delete_user_account
        await delete_user_account(db, db_user_id)
        
        await m.reply(
            f"✅ Пользователь <code>{db_user_id}</code> (Telegram ID/App ID: {target_id}) успешно и полностью удален из базы данных.",
            parse_mode="HTML"
        )
    except Exception as e:
        await m.reply(f"❌ Ошибка при удалении пользователя: {e}")

