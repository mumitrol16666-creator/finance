"""Excel + CSV export of transactions for a chosen period.

Triggered by ``/export`` — the user gets an inline keyboard to pick the period
(today / week / month / all). The result is sent as an in-memory XLSX so we
don't hit disk. CSV is offered as a fallback for users whose phones don't
preview XLSX nicely.

This module is intentionally self-contained: no FSM state, no schema changes.
Reuses ``reports_service`` time helpers so the period bounds respect the
user's timezone.
"""
from __future__ import annotations

import csv
import io
from datetime import datetime, timezone, timedelta
from typing import Iterable

import aiosqlite
from aiogram import F, Router
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.types import BufferedInputFile, CallbackQuery, Message, InlineKeyboardMarkup, InlineKeyboardButton, LabeledPrice
from aiogram.utils.keyboard import InlineKeyboardBuilder

import asyncio
from app.db.repositories.settings_repo import get_lang, get_settings
from app.db.repositories.users_repo import get_free_exports_used, increment_free_export
from app.domain.services.reports_service import day_bounds_utc, month_bounds_utc, iso
from app.domain.time_utils import now_in_user_tz
from app.ui.i18n import t, t_category

router = Router()


def _row_get(row, key: str, index: int, *, default=None):
    """Read a column from an aiosqlite.Row | dict | tuple, with a default."""
    if row is None:
        return default
    if isinstance(row, dict):
        return row.get(key, default)
    try:
        return row[key]
    except (KeyError, IndexError, TypeError):
        pass
    try:
        return row[index]
    except (IndexError, TypeError):
        return default


async def _fetch_rows(
    db: aiosqlite.Connection,
    user_id: int,
    start_iso: str | None,
    end_iso: str | None,
) -> list[tuple]:
    """Pull live (non-deleted) transactions joined with account & category."""
    sql = (
        "SELECT t.id, t.ts, t.type, t.amount, "
        "       COALESCE(a.name, '') AS account, "
        "       COALESCE(c.name, '') AS category, "
        "       COALESCE(c.emoji, '') AS emoji, "
        "       COALESCE(t.note, '') AS note "
        "FROM transactions t "
        "LEFT JOIN accounts a ON a.id = t.account_id "
        "LEFT JOIN categories c ON c.id = t.category_id "
        "WHERE t.user_id=? AND t.deleted_at IS NULL "
    )
    params: list = [user_id]
    if start_iso is not None:
        sql += "AND t.ts>=? "
        params.append(start_iso)
    if end_iso is not None:
        sql += "AND t.ts<? "
        params.append(end_iso)
    sql += "ORDER BY t.ts ASC, t.id ASC"
    cur = await db.execute(sql, params)
    return await cur.fetchall()


def _build_downgrade_preview_png(rows: list, lang: str, currency: str) -> bytes | None:
    """
    Генерирует премиальное превью графиков для бесплатного пользователя,
    чтобы вызвать "эффект потери" и мотивировать к покупке.
    """
    if not rows:
        return None

    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        import io
    except ImportError:
        return None

    # 1. Агрегация данных
    total_income = 0
    total_expense = 0
    expenses_by_cat = {}

    for row in rows:
        try:
            tx_id, ts, ttype, amount, account, category, emoji, note = row
        except Exception:
            continue
        
        is_income = ttype == 'income' 
        amt = abs(int(amount or 0))
        
        if is_income:
            total_income += amt
        else:
            total_expense += amt
            cat_raw = str(category or "").strip()
            if cat_raw:
                cat_name = t_category(cat_raw, lang)
            else:
                cat_name = {
                    "ru": "Другое",
                    "en": "Other",
                    "kk": "Басқа"
                }.get(lang, "Другое")
            expenses_by_cat[cat_name] = expenses_by_cat.get(cat_name, 0) + amt

    # Берем Топ-5 категорий
    top_categories = sorted(expenses_by_cat.items(), key=lambda x: x[1], reverse=True)[:5]
    cat_labels = [item[0] for item in top_categories]
    cat_sizes = [item[1] for item in top_categories]

    # 2. Настройка фигуры и темы
    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6), dpi=150)
    
    # Темная премиальная тема (sleek dark mode)
    bg_color = '#1E1E2E'
    text_color = '#CDD6F4'
    fig.patch.set_facecolor(bg_color)
    
    # Цвета графиков
    color_income = '#A6E3A1'  # Мятный
    color_expense = '#F38BA8' # Рубиновый
    colors_pie = ['#89B4FA', '#F9E2AF', '#F38BA8', '#A6E3A1', '#CBA6F7']

    # --- Левый график: Доходы vs Расходы ---
    ax1.set_facecolor(bg_color)
    
    labels_dynamics = {
        "ru": ['Доходы', 'Расходы'],
        "en": ['Income', 'Expenses'],
        "kk": ['Кірістер', 'Шығыстар']
    }.get(lang, ['Доходы', 'Расходы'])
    
    title_dynamics = {
        "ru": "Денежный поток",
        "en": "Cash Flow",
        "kk": "Ақша ағыны"
    }.get(lang, "Денежный поток")
    
    bars = ax1.bar(labels_dynamics, [total_income, total_expense], color=[color_income, color_expense], width=0.5)
    ax1.set_title(title_dynamics, color=text_color, fontsize=16, pad=20, weight='bold')
    ax1.tick_params(colors=text_color, labelsize=12)
    
    # Убираем рамки
    for spine in ax1.spines.values():
        spine.set_visible(False)
    ax1.grid(axis='y', color='#313244', linestyle='--', alpha=0.7)

    # --- Правый график: Топ-5 категорий расходов ---
    title_pie = {
        "ru": "Топ расходов",
        "en": "Top Expenses",
        "kk": "Ең көп шығыстар"
    }.get(lang, "Топ расходов")
    
    if cat_sizes:
        _, texts, autotexts = ax2.pie(
            cat_sizes, 
            labels=cat_labels, 
            colors=colors_pie,
            autopct='%1.1f%%', 
            startangle=140,
            textprops={'color': text_color, 'fontsize': 12},
            wedgeprops={'edgecolor': bg_color, 'linewidth': 2}
        )
        for autotext in autotexts:
            autotext.set_color(bg_color)
            autotext.set_weight('bold')
        ax2.set_title(title_pie, color=text_color, fontsize=16, pad=20, weight='bold')
    else:
        ax2.set_facecolor(bg_color)
        ax2.text(0.5, 0.5, {
            "ru": "Нет расходов за период",
            "en": "No expenses",
            "kk": "Шығыстар жоқ"
        }.get(lang, "Нет расходов"), color=text_color, ha='center', va='center', fontsize=14)
        for spine in ax2.spines.values():
            spine.set_visible(False)
        ax2.set_xticks([])
        ax2.set_yticks([])

    # --- 3. Водяной знак ---
    watermark_text = {
        "ru": "ПРОФЕССИОНАЛЬНАЯ АНАЛИТИКА · FINANCEBOT",
        "en": "PROFESSIONAL ANALYTICS · FINANCEBOT",
        "kk": "КӘСІБИ ТАЛДАУ · FINANCEBOT"
    }.get(lang, "ПРОФЕССИОНАЛЬНАЯ АНАЛИТИКА · FINANCEBOT")
    
    fig.text(0.5, 0.5, watermark_text, 
             fontsize=32, color='white', alpha=0.04, 
             ha='center', va='center', rotation=15, weight='bold')

    # 4. Сохранение в байты
    buf = io.BytesIO()
    plt.tight_layout()
    plt.savefig(buf, format='png', facecolor=fig.get_facecolor(), bbox_inches='tight', dpi=150)
    plt.close(fig)
    buf.seek(0)
    
    return buf.getvalue()


def format_localized_datetime(dt, lang: str) -> str:
    months = {
        "ru": ["января", "февраля", "марта", "апреля", "мая", "июня", "июля", "августа", "сентября", "октября", "ноября", "декабря"],
        "en": ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"],
        "kk": ["қаңтар", "ақпан", "наурыз", "сәуір", "мамыр", "маусым", "шілде", "тамыз", "қыркүйек", "қазан", "қараша", "желтоқсан"]
    }
    lang_months = months.get(lang, months["ru"])
    month_name = lang_months[dt.month - 1]
    
    time_str = dt.strftime("%H:%M")
    if lang == "ru":
        return f"{dt.day} {month_name} {dt.year} г. {time_str}"
    elif lang == "kk":
        return f"{dt.year} жылғы {dt.day} {month_name}, {time_str}"
    else: # en
        return f"{month_name} {dt.day}, {dt.year} {time_str}"


def format_localized_month(month_str: str, lang: str) -> str:
    try:
        parts = month_str.split("-")
        if len(parts) != 2:
            return month_str
        year, month_idx = int(parts[0]), int(parts[1])
        months = {
            "ru": ["Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"],
            "en": ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"],
            "kk": ["Қаңтар", "Ақпан", "Наурыз", "Сәуір", "Мамыр", "Маусым", "Шілде", "Тамыз", "Қыркүйек", "Қазан", "Қараша", "Желтоқсан"]
        }
        lang_months = months.get(lang, months["ru"])
        month_name = lang_months[month_idx - 1]
        return f"{month_name} {year}"
    except Exception:
        return month_str


def _build_xlsx(
    rows: Iterable[tuple],
    lang: str,
    currency: str,
    user_id: int,
    metrics: dict,
    profile: dict,
    priority_insights: dict,
    latest_rec: dict | None,
    tz_name: str = "Asia/Aqtobe",
    accounts_data: list = None,
    budgets_data: list = None,
    liabilities_data: list = None,
    recurring_data: list = None,
    planned_data: list = None,
    all_insights: list = None
) -> bytes | None:
    """Render rows into a premium, app-like 4-sheet Excel workbook.
    
    Aesthetics:
    - Minimalist Apple/Stripe-inspired Slate & Emerald theme.
    - Gridlines disabled on Summary and Analytics sheets for a dashboard feel.
    - Large row heights (22-35px) for spacious typography.
    - Explicit column widths to completely avoid text truncation (### or clipped text).
    - Localized dates in user's timezone (YYYY-MM-DD HH:MM).
    - Fully localized transaction types (Доход / Расход / Перевод).
    """
    try:
        import collections
        import io
        import calendar
        from datetime import datetime, timezone
        from zoneinfo import ZoneInfo
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    except ImportError:
        return None

    # Localized labels
    all_labels = {
        "ru": {
            "title_summary": "ФИНАНСОВЫЙ ПУЛЬС • ГЛАВНЫЙ ДАШБОРД",
            "health_title": "ФИНАНСОВЫЙ ИНДЕКС",
            "runway_title": "ЗАПАС ПРОЧНОСТИ",
            "burn_title": "СУТОЧНЫЙ РАСХОД",
            "main_risk": "КЛЮЧЕВОЙ ФИНАНСОВЫЙ РИСК",
            "ai_rec": "AI РЕКОМЕНДАЦИЯ КОУЧА",
            "behavioral": "ПОВЕДЕНЧЕСКИЙ ПРОФИЛЬ",
            "monthly_title": "Динамика доходов и расходов",
            "col_month": "Месяц",
            "col_income": "Доходы",
            "col_expense": "Расходы",
            "col_net": "Чистый доход",
            "col_savings": "Сбережения %",
            "category_title": "Структура расходов по категориям",
            "col_category": "Категория",
            "col_spent": "Потрачено",
            "col_share": "Доля расходов",
            "total": "Итого",
            "trends_title": "Сравнение трендов (vs пред. месяц)",
            "income_trend": "Изменение доходов",
            "expense_trend": "Изменение расходов",
            "savings_trend": "Коэффициент сбережений",
            "prediction_title": "Прогноз на конец месяца",
            "projected_spending": "Ожидаемые расходы к концу месяца",
            "anomalies_title": "Аномалии и предупреждения",
            "no_anomalies": "Аномалий или повышенных рисков не обнаружено.",
            "sheet_dashboard": "Главная сводка",
            "sheet_analytics": "Аналитика",
            "sheet_schedules": "Долги и расписание",
            "sheet_transactions": "История операций",
            "liabilities_title": "Кредиты и долги",
            "recurring_title": "Регулярные платежи и подписки",
            "planned_title": "Запланированные операции",
            "ai_analytic_insight": "AI АНАЛИЗ КАТЕГОРИЙ",
            "raw_headers": ["ID", "Дата и время", "Тип", "Сумма", "Валюта", "Счёт", "Категория", "Комментарий"]
        },
        "en": {
            "title_summary": "FINANCIAL PULSE • EXECUTIVE SUMMARY",
            "health_title": "FINANCIAL HEALTH INDEX",
            "runway_title": "RESERVES RUNWAY",
            "burn_title": "DAILY BURN RATE",
            "main_risk": "MAIN FINANCIAL RISK",
            "ai_rec": "AI COACH RECOMMENDATION",
            "behavioral": "BEHAVIORAL PROFILE",
            "monthly_title": "Monthly Cashflow Dynamics",
            "col_month": "Month",
            "col_income": "Income",
            "col_expense": "Expenses",
            "col_net": "Net Income",
            "col_savings": "Savings %",
            "category_title": "Expenses Structure by Category",
            "col_category": "Category",
            "col_spent": "Spent",
            "col_share": "Expense Share",
            "total": "Total",
            "trends_title": "Trend Comparison (vs last month)",
            "income_trend": "Income change",
            "expense_trend": "Expenses change",
            "savings_trend": "Savings rate change",
            "prediction_title": "Month-End Prediction",
            "projected_spending": "Projected month-end spending",
            "anomalies_title": "Anomaly Alerts",
            "no_anomalies": "No anomalies or high risks detected.",
            "sheet_dashboard": "Executive Summary",
            "sheet_analytics": "Analytics",
            "sheet_schedules": "Debts & Schedules",
            "sheet_transactions": "Transaction History",
            "liabilities_title": "Debts & Liabilities",
            "recurring_title": "Recurring Payments & Subs",
            "planned_title": "Planned Transactions",
            "ai_analytic_insight": "AI CATEGORY ANALYSIS",
            "raw_headers": ["ID", "Date & Time", "Type", "Amount", "Currency", "Account", "Category", "Note"]
        },
        "kk": {
            "title_summary": "ҚАРЖЫЛЫҚ ТАМЫР ЛҮПІЛІ • НЕГІЗГІ ДАШБОРД",
            "health_title": "ҚАРЖЫЛЫҚ ИНДЕКС",
            "runway_title": "ҚАУІПСІЗДІК ҚОРЫ",
            "burn_title": "КҮНДЕЛІКТІ ШЫҒЫН",
            "main_risk": "НЕГІЗГІ ҚАРЖЫЛЫҚ ҚАУІП",
            "ai_rec": "AI КОУЧ ҰСЫНЫСЫ",
            "behavioral": "МІНЕЗ-ҚҰЛЫҚ ПРОФИЛІ",
            "monthly_title": "Кірістер мен шығыстар динамикасы",
            "col_month": "Ай",
            "col_income": "Кірістер",
            "col_expense": "Шығыстар",
            "col_net": "Таза кіріс",
            "col_savings": "Жинақ %",
            "category_title": "Санаттар бойынша шығыстар",
            "col_category": "Санат",
            "col_spent": "Жұмсалды",
            "col_share": "Шығыс үлесі",
            "total": "Жиынтығы",
            "trends_title": "Трендтерді салыстыру (өткен аймен)",
            "income_trend": "Кірістердің өзгеруі",
            "expense_trend": "Шығыстардың өзгеруі",
            "savings_trend": "Жинақ коэффициенті",
            "prediction_title": "Ай соңына болжам",
            "projected_spending": "Айдың соңына күтілетін шығыстар",
            "anomalies_title": "Аномалиялар мен ескертулер",
            "no_anomalies": "Аномалиялар немесе жоғары қауіптер табылған жоқ.",
            "sheet_dashboard": "Негізгі жиынтық",
            "sheet_analytics": "Талдау",
            "sheet_schedules": "Борыштар мен кесте",
            "sheet_transactions": "Операциялар тарихы",
            "liabilities_title": "Несиелер мен борыштар",
            "recurring_title": "Тұрақты төлемдер",
            "planned_title": "Жоспарланған операциялар",
            "ai_analytic_insight": "AI САНАТТАРДЫ ТАЛДАУ",
            "raw_headers": ["ID", "Күні мен уақыты", "Түрі", "Сома", "Валюта", "Шот", "Санат", "Түсініктеме"]
        }
    }
    L = all_labels.get(lang, all_labels["ru"])

    # 1. Process Data / Aggregations
    total_income = 0
    total_expense = 0
    monthly_data = collections.defaultdict(lambda: {"income": 0, "expense": 0})
    category_data = collections.defaultdict(int)

    rows_list = list(rows)
    for r in rows_list:
        tx_id, ts, ttype, amount, account, category, emoji, note = r
        val = abs(int(amount or 0))
        
        try:
            dt = datetime.fromisoformat(ts)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            local_dt = dt.astimezone(ZoneInfo(tz_name))
            month_str = local_dt.strftime("%Y-%m")
        except Exception:
            month_str = str(ts)[:7] if ts and len(str(ts)) >= 7 else "Unknown"
        
        if ttype == "income":
            total_income += val
            monthly_data[month_str]["income"] += val
        elif ttype == "expense":
            total_expense += val
            monthly_data[month_str]["expense"] += val
            translated_cat = t_category(category, lang)
            other_lbl = {"ru": "Прочее", "en": "Other", "kk": "Басқа"}.get(lang, "Other")
            cat_display = f"{emoji} {translated_cat}".strip() if emoji or translated_cat else other_lbl
            category_data[cat_display] += val

    wb = Workbook()
    
    font_main_title = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    font_sec_hdr = Font(name="Segoe UI", size=11, bold=True, color="1E293B")
    font_tbl_hdr = Font(name="Segoe UI", size=10, bold=True, color="475569")
    font_bold = Font(name="Segoe UI", size=10, bold=True, color="0F172A")
    font_data = Font(name="Segoe UI", size=10, color="334155")
    font_subtext = Font(name="Segoe UI", size=9, color="64748B")
    font_mono = Font(name="Consolas", size=10, color="0F172A")
    
    font_card_val = Font(name="Segoe UI", size=22, bold=True, color="0F172A")
    
    fill_health = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    fill_runway = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
    fill_burn = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    font_card_title_1 = Font(name="Segoe UI", size=9, bold=True, color="1E40AF")
    font_card_title_2 = Font(name="Segoe UI", size=9, bold=True, color="065F46")
    font_card_title_3 = Font(name="Segoe UI", size=9, bold=True, color="334155")
    
    fill_header_banner = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    fill_card = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_tbl_hdr = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    fill_accent_green = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
    
    thin_gray = Side(style='thin', color='E2E8F0')
    border_all = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    border_bottom = Border(bottom=Side(style='thin', color='E2E8F0'))
    border_double_bottom = Border(top=Side(style='thin', color='94A3B8'), bottom=Side(style='double', color='0F172A'))

    # ==========================================
    # SHEET 1: EXECUTIVE SUMMARY (App-like Dashboard)
    # ==========================================
    ws_sum = wb.active
    ws_sum.title = L["sheet_dashboard"]
    ws_sum.views.sheetView[0].showGridLines = False

    ws_sum.merge_cells("A1:K1")
    ws_sum["A1"] = L["title_summary"]
    ws_sum["A1"].font = font_main_title
    ws_sum["A1"].fill = fill_header_banner
    ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 42

    sav_pct = metrics.get("savings_rate", {}).get("savings_rate_pct", 0)
    sav_score = min(40, max(0, int(sav_pct * 1.33)))
    runway = metrics.get("runway_days", 0)
    if runway == float('inf') or runway is None:
        runway_val = 99
        runway_score = 30
    else:
        runway_val = int(runway)
        runway_score = min(30, max(0, int(runway / 3.0)))
    discipline = profile.get("discipline_score", 100)
    discipline_score = min(30, max(0, int(discipline * 0.3)))
    health_score = sav_score + runway_score + discipline_score

    ws_sum.merge_cells("A3:C3")
    ws_sum.merge_cells("A4:C4")
    ws_sum.merge_cells("A5:C5")
    ws_sum["A3"] = L["health_title"]
    ws_sum["A3"].font = font_card_title_1
    ws_sum["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sum["A4"] = f"{health_score} / 100"
    ws_sum["A4"].font = font_card_val
    ws_sum["A4"].alignment = Alignment(horizontal="center", vertical="center")
    
    score_trend = {
        "ru": "Высокая дисциплина" if discipline > 80 else "Стабильный",
        "en": "High discipline" if discipline > 80 else "Stable",
        "kk": "Жоғары тәртіп" if discipline > 80 else "Тұрақты"
    }.get(lang, "High discipline" if discipline > 80 else "Stable")
    ws_sum["A5"] = score_trend
    ws_sum["A5"].font = font_subtext
    ws_sum["A5"].alignment = Alignment(horizontal="center", vertical="center")

    ws_sum.merge_cells("E3:G3")
    ws_sum.merge_cells("E4:G4")
    ws_sum.merge_cells("E5:G5")
    ws_sum["E3"] = L["runway_title"]
    ws_sum["E3"].font = font_card_title_2
    ws_sum["E3"].alignment = Alignment(horizontal="center", vertical="center")
    ws_sum["E4"] = f"{runway_val} дней" if lang == "ru" else (f"{runway_val} күн" if lang == "kk" else f"{runway_val} Days")
    ws_sum["E4"].font = font_card_val
    ws_sum["E4"].alignment = Alignment(horizontal="center", vertical="center")
    
    runway_desc = {
        "ru": "Резервы ликвидности",
        "en": "Liquidity reserves",
        "kk": "Ликвидтілік резервтері"
    }.get(lang, "Liquidity reserves")
    ws_sum["E5"] = runway_desc
    ws_sum["E5"].font = font_subtext
    ws_sum["E5"].alignment = Alignment(horizontal="center", vertical="center")

    ws_sum.merge_cells("I3:K3")
    ws_sum.merge_cells("I4:K4")
    ws_sum.merge_cells("I5:K5")
    ws_sum["I3"] = L["burn_title"]
    ws_sum["I3"].font = font_card_title_3
    ws_sum["I3"].alignment = Alignment(horizontal="center", vertical="center")
    
    burn_daily = metrics.get("burn_rate", {}).get("daily_burn_rate", 0)
    ws_sum["I4"] = f"{burn_daily} {currency}"
    ws_sum["I4"].font = font_card_val
    ws_sum["I4"].alignment = Alignment(horizontal="center", vertical="center")
    
    burn_trend_pct = metrics.get("burn_rate", {}).get("trend_pct", 0.0)
    trend_sign = "+" if burn_trend_pct > 0 else ""
    burn_desc = {
        "ru": f"{trend_sign}{burn_trend_pct}% vs пред. период",
        "en": f"{trend_sign}{burn_trend_pct}% vs last period",
        "kk": f"{trend_sign}{burn_trend_pct}% өткен кезеңмен"
    }.get(lang, f"{trend_sign}{burn_trend_pct}% vs last period")
    ws_sum["I5"] = burn_desc
    ws_sum["I5"].font = font_subtext
    ws_sum["I5"].alignment = Alignment(horizontal="center", vertical="center")

    for c in ["A", "B", "C"]:
        for r in [3, 4, 5]:
            ws_sum[f"{c}{r}"].fill = fill_health
            ws_sum[f"{c}{r}"].border = Border(
                top=thin_gray if r == 3 else None,
                bottom=thin_gray if r == 5 else None,
                left=thin_gray if c == "A" else None,
                right=thin_gray if c == "C" else None
            )
            
    for c in ["E", "F", "G"]:
        for r in [3, 4, 5]:
            ws_sum[f"{c}{r}"].fill = fill_runway
            ws_sum[f"{c}{r}"].border = Border(
                top=thin_gray if r == 3 else None,
                bottom=thin_gray if r == 5 else None,
                left=thin_gray if c == "E" else None,
                right=thin_gray if c == "G" else None
            )
            
    for c in ["I", "J", "K"]:
        for r in [3, 4, 5]:
            ws_sum[f"{c}{r}"].fill = fill_burn
            ws_sum[f"{c}{r}"].border = Border(
                top=thin_gray if r == 3 else None,
                bottom=thin_gray if r == 5 else None,
                left=thin_gray if c == "I" else None,
                right=thin_gray if c == "K" else None
            )

    ws_sum.row_dimensions[3].height = 20
    ws_sum.row_dimensions[4].height = 32
    ws_sum.row_dimensions[5].height = 20

    ws_sum["A7"] = L["main_risk"]
    ws_sum["A7"].font = font_sec_hdr
    ws_sum.row_dimensions[7].height = 28
    
    main_problem_text = {
        "ru": "Рисков не обнаружено",
        "en": "No risks detected",
        "kk": "Қауіп табылған жоқ"
    }.get(lang, "No risks detected")
    mp = priority_insights.get("main_problem")
    if mp:
        main_problem_text = mp.get("text", main_problem_text)
    
    ws_sum.merge_cells("A8:K8")
    ws_sum["A8"] = f" ⚠️  {main_problem_text}"
    ws_sum["A8"].font = font_bold
    ws_sum["A8"].alignment = Alignment(wrap_text=True, vertical="center", indent=1)
    ws_sum["A8"].border = border_all
    ws_sum["A8"].fill = fill_card
    ws_sum.row_dimensions[8].height = 30

    ws_sum["A10"] = L["ai_rec"]
    ws_sum["A10"].font = font_sec_hdr
    ws_sum.row_dimensions[10].height = 28
    
    rec_text = {
        "ru": "Продолжайте удерживать текущие лимиты расходов.",
        "en": "Keep maintaining current spending limits.",
        "kk": "Шығындар лимиттерін ұстап тұруды жалғастырыңыз."
    }.get(lang, "Keep maintaining current spending limits.")
    if latest_rec:
        rec_text = latest_rec.get("text", rec_text)
        
    ws_sum.merge_cells("A11:K11")
    ws_sum["A11"] = f" 💡  {rec_text}"
    ws_sum["A11"].font = font_data
    ws_sum["A11"].alignment = Alignment(wrap_text=True, vertical="center", indent=1)
    ws_sum["A11"].border = border_all
    ws_sum["A11"].fill = fill_accent_green
    ws_sum.row_dimensions[11].height = 30

    ws_sum["A13"] = L["behavioral"]
    ws_sum["A13"].font = font_sec_hdr
    ws_sum.row_dimensions[13].height = 28
    
    beh_summary = profile.get("behavioral_summary", "накапливаем статистику")
    if beh_summary == "накапливаем статистику":
        beh_summary = {
            "ru": "накапливаем статистику",
            "en": "collecting statistics",
            "kk": "статистика жиналуда"
        }.get(lang, "collecting statistics")
        
    ws_sum.merge_cells("A14:K14")
    ws_sum["A14"] = f" 👤  {beh_summary}"
    ws_sum["A14"].font = font_data
    ws_sum["A14"].alignment = Alignment(wrap_text=True, vertical="center", indent=1)
    ws_sum["A14"].border = border_all
    ws_sum["A14"].fill = fill_card
    ws_sum.row_dimensions[14].height = 30

    # Accounts & Balances Section (A16 onwards)
    if accounts_data:
        ws_sum.cell(row=16, column=1, value={"ru": "💳 Счета и текущие балансы", "en": "💳 Accounts & Balances", "kk": "💳 Шоттар және ағымдағы баланстар"}.get(lang)).font = font_sec_hdr
        ws_sum.row_dimensions[16].height = 28
        
        headers_acc = {
            "ru": ["Счёт", "", "", "Тип", "Валюта", "Стартовый", "Текущий", "", "Изменение", "Рост %"],
            "en": ["Account", "", "", "Type", "Currency", "Starting", "Current", "", "Change", "Growth %"],
            "kk": ["Шот", "", "", "Түрі", "Валюта", "Бастапқы", "Ағымдағы", "", "Өзгеруі", "Өсу %"]
        }.get(lang)
        
        ws_sum.merge_cells("A17:C17")
        ws_sum["A17"] = headers_acc[0]
        ws_sum.cell(row=17, column=4, value=headers_acc[3])
        ws_sum.cell(row=17, column=5, value=headers_acc[4])
        ws_sum.cell(row=17, column=6, value=headers_acc[5])
        ws_sum.cell(row=17, column=7, value=headers_acc[6])
        ws_sum.cell(row=17, column=9, value=headers_acc[8])
        ws_sum.cell(row=17, column=10, value=headers_acc[9])
        
        for col_idx in [1, 2, 3, 4, 5, 6, 7, 9, 10]:
            c_cell = ws_sum.cell(row=17, column=col_idx)
            c_cell.font = font_tbl_hdr
            c_cell.fill = fill_tbl_hdr
            c_cell.alignment = Alignment(horizontal="center", vertical="center")
            c_cell.border = border_all
        ws_sum.row_dimensions[17].height = 24
        
        a_idx = 18
        for row in accounts_data:
            # acc_id, name, balance, starting_balance, currency, is_saving
            acc_id, name, balance, starting_balance, curr, is_saving = row
            starting_balance = int(starting_balance or 0)
            balance = int(balance or 0)
            
            type_str = {
                "ru": "Накопительный" if is_saving else "Обычный",
                "en": "Saving" if is_saving else "Regular",
                "kk": "Жинақтаушы" if is_saving else "Қалыпты"
            }.get(lang)
            
            diff = balance - starting_balance
            growth_pct = (diff / starting_balance) if starting_balance > 0 else 0.0
            
            ws_sum.merge_cells(start_row=a_idx, start_column=1, end_row=a_idx, end_column=3)
            ws_sum.cell(row=a_idx, column=1, value=name).alignment = Alignment(horizontal="left", indent=1)
            ws_sum.cell(row=a_idx, column=4, value=type_str).alignment = Alignment(horizontal="center")
            ws_sum.cell(row=a_idx, column=5, value=curr).alignment = Alignment(horizontal="center")
            ws_sum.cell(row=a_idx, column=6, value=starting_balance).number_format = '#,##0'
            ws_sum.cell(row=a_idx, column=7, value=balance).number_format = '#,##0'
            ws_sum.cell(row=a_idx, column=9, value=diff).number_format = '#,##0'
            ws_sum.cell(row=a_idx, column=10, value=growth_pct).number_format = '0.0%'
            
            for col_idx in [1, 2, 3, 4, 5, 6, 7, 9, 10]:
                cell = ws_sum.cell(row=a_idx, column=col_idx)
                cell.font = font_data
                cell.border = border_all
                if a_idx % 2 == 1:
                    cell.fill = fill_zebra
            ws_sum.row_dimensions[a_idx].height = 20
            a_idx += 1
            
        ws_sum.merge_cells(start_row=a_idx, start_column=1, end_row=a_idx, end_column=3)
        ws_sum.cell(row=a_idx, column=1, value=L["total"]).font = font_bold
        ws_sum.cell(row=a_idx, column=1).alignment = Alignment(horizontal="center")
        
        tot_starting = sum(int(r[3] or 0) for r in accounts_data)
        tot_current = sum(int(r[2] or 0) for r in accounts_data)
        tot_diff = tot_current - tot_starting
        tot_growth = (tot_diff / tot_starting) if tot_starting > 0 else 0.0
        
        ws_sum.cell(row=a_idx, column=6, value=tot_starting).font = font_bold
        ws_sum.cell(row=a_idx, column=6).number_format = '#,##0'
        ws_sum.cell(row=a_idx, column=7, value=tot_current).font = font_bold
        ws_sum.cell(row=a_idx, column=7).number_format = '#,##0'
        
        ws_sum.cell(row=a_idx, column=9, value=tot_diff).font = font_bold
        ws_sum.cell(row=a_idx, column=9).number_format = '#,##0'
        ws_sum.cell(row=a_idx, column=10, value=tot_growth).font = font_bold
        ws_sum.cell(row=a_idx, column=10).number_format = '0.0%'
        
        for col_idx in [1, 2, 3, 4, 5, 6, 7, 9, 10]:
            ws_sum.cell(row=a_idx, column=col_idx).border = border_double_bottom
        ws_sum.row_dimensions[a_idx].height = 22

    widths_sum = {"A": 15, "B": 15, "C": 15, "D": 15, "E": 15, "F": 15, "G": 15, "H": 5, "I": 15, "J": 15, "K": 15}
    for col, w in widths_sum.items():
        ws_sum.column_dimensions[col].width = w

    # ==========================================
    # SHEET 2: ANALYTICS (Structured Data Blocks)
    # ==========================================
    ws_an = wb.create_sheet(title=L["sheet_analytics"])
    ws_an.views.sheetView[0].showGridLines = False

    # Title Dynamic Cashflow
    ws_an["A1"] = L["monthly_title"]
    ws_an["A1"].font = font_sec_hdr
    ws_an.row_dimensions[1].height = 30

    headers_m = [L["col_month"], L["col_income"], L["col_expense"], L["col_net"], L["col_savings"]]
    for c_idx, h in enumerate(headers_m, start=1):
        cell = ws_an.cell(row=2, column=c_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_all
    ws_an.row_dimensions[2].height = 24

    sorted_months = sorted(list(monthly_data.keys()))
    row_idx = 3
    for m_str in sorted_months:
        inc = monthly_data[m_str]["income"]
        exp = monthly_data[m_str]["expense"]
        net_val = inc - exp
        sav_val = net_val / inc if inc > 0 else 0.0
        
        ws_an.cell(row=row_idx, column=1, value=format_localized_month(m_str, lang)).alignment = Alignment(horizontal="center")
        ws_an.cell(row=row_idx, column=2, value=inc)
        ws_an.cell(row=row_idx, column=3, value=exp)
        ws_an.cell(row=row_idx, column=4, value=net_val)
        ws_an.cell(row=row_idx, column=5, value=sav_val)
        
        ws_an.cell(row=row_idx, column=2).number_format = '#,##0'
        ws_an.cell(row=row_idx, column=3).number_format = '#,##0'
        ws_an.cell(row=row_idx, column=4).number_format = '#,##0'
        ws_an.cell(row=row_idx, column=5).number_format = '0.0%'
        
        for c in range(1, 6):
            cell = ws_an.cell(row=row_idx, column=c)
            cell.font = font_data
            cell.border = border_all
            if row_idx % 2 == 1:
                cell.fill = fill_zebra
        ws_an.row_dimensions[row_idx].height = 20
        row_idx += 1

    # Monthly Total
    monthly_net_total = total_income - total_expense
    monthly_sav_total = monthly_net_total / total_income if total_income > 0 else 0.0

    ws_an.cell(row=row_idx, column=1, value=L["total"]).font = font_bold
    ws_an.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")
    ws_an.cell(row=row_idx, column=2, value=total_income).font = font_bold
    ws_an.cell(row=row_idx, column=3, value=total_expense).font = font_bold
    ws_an.cell(row=row_idx, column=4, value=monthly_net_total).font = font_bold
    ws_an.cell(row=row_idx, column=5, value=monthly_sav_total).font = font_bold
    
    ws_an.cell(row=row_idx, column=2).number_format = '#,##0'
    ws_an.cell(row=row_idx, column=3).number_format = '#,##0'
    ws_an.cell(row=row_idx, column=4).number_format = '#,##0'
    ws_an.cell(row=row_idx, column=5).number_format = '0.0%'
    
    for c in range(1, 6):
        ws_an.cell(row=row_idx, column=c).border = border_double_bottom
    ws_an.row_dimensions[row_idx].height = 22
    
    last_month_row = row_idx

    # Trend Comparison Block (Left Side)
    start_trend_row = last_month_row + 3
    ws_an.cell(row=start_trend_row, column=1, value=L["trends_title"]).font = font_sec_hdr
    ws_an.row_dimensions[start_trend_row].height = 28
    
    # Calculate side-by-side month comparison
    if len(sorted_months) >= 2:
        m_curr, m_prev = sorted_months[-1], sorted_months[-2]
        inc_c = monthly_data[m_curr]["income"]
        inc_p = monthly_data[m_prev]["income"]
        exp_c = monthly_data[m_curr]["expense"]
        exp_p = monthly_data[m_prev]["expense"]
    elif len(sorted_months) == 1:
        m_curr = sorted_months[0]
        m_prev = None
        inc_c = monthly_data[m_curr]["income"]
        inc_p = 0
        exp_c = monthly_data[m_curr]["expense"]
        exp_p = 0
    else:
        m_curr = None
        m_prev = None
        inc_c = 0
        inc_p = 0
        exp_c = 0
        exp_p = 0

    net_c = inc_c - exp_c
    net_p = inc_p - exp_p
    sav_c = (net_c / inc_c) if inc_c > 0 else 0.0
    sav_p = (net_p / inc_p) if inc_p > 0 else 0.0

    prev_hdr = format_localized_month(m_prev, lang) if m_prev else "-"
    curr_hdr = format_localized_month(m_curr, lang) if m_curr else "-"

    comp_headers = {
        "ru": ["Показатель", prev_hdr, curr_hdr, "Изменение"],
        "en": ["Metric", prev_hdr, curr_hdr, "Change"],
        "kk": ["Көрсеткіш", prev_hdr, curr_hdr, "Өзгеруі"]
    }.get(lang, ["Metric", prev_hdr, curr_hdr, "Change"])

    for c_idx, h in enumerate(comp_headers, start=1):
        cell = ws_an.cell(row=start_trend_row + 1, column=c_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_all
    ws_an.row_dimensions[start_trend_row + 1].height = 22

    def get_pct_change_str(curr, prev, higher_is_bad=False):
        if prev == 0:
            if curr == 0:
                return "0.0%"
            return "+100.0% 📈" if not higher_is_bad else "+100.0% 📈 ⚠️"
        val = ((curr - prev) / prev) * 100.0
        sign = "+" if val > 0 else ""
        emoji = ""
        if val > 0:
            emoji = " 📈 ⚠️" if higher_is_bad else " 📈"
        elif val < 0:
            emoji = " 📉" if higher_is_bad else " 📉 ⚠️"
        return f"{sign}{val:.1f}%{emoji}"

    def get_net_change_str(curr, prev):
        diff = curr - prev
        sign = "+" if diff > 0 else ""
        emoji = " 🚀" if diff > 0 else " 📉"
        return f"{sign}{diff:+,} {currency}{emoji}".replace("+-", "-").replace("++", "+")

    pct_inc = get_pct_change_str(inc_c, inc_p, False)
    pct_exp = get_pct_change_str(exp_c, exp_p, True)
    net_change = get_net_change_str(net_c, net_p)
    sav_diff_pct = (sav_c - sav_p) * 100.0
    sign_sav = "+" if sav_diff_pct > 0 else ""
    emoji_sav = " 📈" if sav_diff_pct > 0 else " 📉 ⚠️"
    sav_change = f"{sign_sav}{sav_diff_pct:+.1f}%{emoji_sav}".replace("+-", "-").replace("++", "+")

    row_labels = {
        "ru": ["Доходы", "Расходы", "Чистый доход", "Норма сбережений"],
        "en": ["Income", "Expenses", "Net Income", "Savings Rate"],
        "kk": ["Кірістер", "Шығыстар", "Таза кіріс", "Жинақ нормасы"]
    }.get(lang, ["Income", "Expenses", "Net Income", "Savings Rate"])

    comp_rows = [
        (row_labels[0], inc_p, inc_c, pct_inc, False),
        (row_labels[1], exp_p, exp_c, pct_exp, False),
        (row_labels[2], net_p, net_c, net_change, True),
        (row_labels[3], sav_p, sav_c, sav_change, False, "0.0%")
    ]

    t_row = start_trend_row + 2
    for r_idx, item in enumerate(comp_rows):
        label, p_val, c_val, change_str, is_raw_text, *custom_fmt = item
        ws_an.cell(row=t_row, column=1, value=label).font = font_data
        ws_an.cell(row=t_row, column=1).border = border_all
        
        cell_p = ws_an.cell(row=t_row, column=2)
        if not is_raw_text:
            cell_p.value = p_val
            cell_p.number_format = custom_fmt[0] if custom_fmt else '#,##0'
        else:
            cell_p.value = f"{p_val:+,}".replace("+-", "-").replace("++", "+")
        cell_p.font = font_data
        cell_p.border = border_all
        cell_p.alignment = Alignment(horizontal="right")

        cell_c = ws_an.cell(row=t_row, column=3)
        if not is_raw_text:
            cell_c.value = c_val
            cell_c.number_format = custom_fmt[0] if custom_fmt else '#,##0'
        else:
            cell_c.value = f"{c_val:+,}".replace("+-", "-").replace("++", "+")
        cell_c.font = font_data
        cell_c.border = border_all
        cell_c.alignment = Alignment(horizontal="right")

        cell_d = ws_an.cell(row=t_row, column=4, value=change_str)
        cell_d.font = font_bold
        cell_d.border = border_all
        cell_d.alignment = Alignment(horizontal="center")
        
        if t_row % 2 == 1:
            for col in range(1, 5):
                ws_an.cell(row=t_row, column=col).fill = fill_zebra
        ws_an.row_dimensions[t_row].height = 20
        t_row += 1

    end_left_row = t_row

    # ==========================================
    # RIGHT SIDE (Columns H:L)
    # ==========================================
    # 1. Category breakdown
    ws_an["H1"] = L["category_title"]
    ws_an["H1"].font = font_sec_hdr
    ws_an.row_dimensions[1].height = 30

    headers_c = [L["col_category"], L["col_spent"], L["col_share"]]
    for c_idx, h in enumerate(headers_c, start=8):  # H, I, J
        cell = ws_an.cell(row=2, column=c_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_all
    ws_an.row_dimensions[2].height = 24

    sorted_cats = sorted(category_data.items(), key=lambda x: x[1], reverse=True)
    c_row_idx = 3
    total_category_spent = sum(category_data.values())

    if sorted_cats:
        for cat_name, spent in sorted_cats:
            ws_an.cell(row=c_row_idx, column=8, value=cat_name)
            ws_an.cell(row=c_row_idx, column=9, value=spent)
            
            share_val = spent / total_category_spent if total_category_spent > 0 else 0.0
            num_blocks = int(round(share_val * 10))
            bar_str = "█" * num_blocks + "░" * (10 - num_blocks)
            ws_an.cell(row=c_row_idx, column=10, value=f"{bar_str}  {share_val * 100:.1f}%")
            
            ws_an.cell(row=c_row_idx, column=9).number_format = '#,##0'
            ws_an.cell(row=c_row_idx, column=10).alignment = Alignment(horizontal="left")
            
            for c in range(8, 11):
                cell = ws_an.cell(row=c_row_idx, column=c)
                cell.font = font_data
                cell.border = border_all
                if c_row_idx % 2 == 1:
                    cell.fill = fill_zebra
            ws_an.row_dimensions[c_row_idx].height = 20
            c_row_idx += 1

        ws_an.cell(row=c_row_idx, column=8, value=L["total"]).font = font_bold
        ws_an.cell(row=c_row_idx, column=8).alignment = Alignment(horizontal="center")
        ws_an.cell(row=c_row_idx, column=9, value=total_category_spent).font = font_bold
        ws_an.cell(row=c_row_idx, column=10, value="██████████  100%").font = font_bold
        ws_an.cell(row=c_row_idx, column=9).number_format = '#,##0'
        
        for c in range(8, 11):
            ws_an.cell(row=c_row_idx, column=c).border = border_double_bottom
        ws_an.row_dimensions[c_row_idx].height = 22
    else:
        ws_an.cell(row=3, column=8, value="-").alignment = Alignment(horizontal="center")
        ws_an.cell(row=3, column=9, value=0)
        ws_an.cell(row=3, column=10, value="░░░░░░░░░░  0%")
        c_row_idx = 3

    last_cat_row = c_row_idx

    # 2. Budgets & Limits (Columns H:L)
    start_bud_row = last_cat_row + 2
    budgets_title_str = {
        "ru": "📌 Лимиты и бюджеты категорий",
        "en": "📌 Category Budgets & Limits",
        "kk": "📌 Санаттар лимиттері мен бюджеттері"
    }.get(lang, "📌 Лимиты и бюджеты категорий")
    ws_an.cell(row=start_bud_row, column=8, value=budgets_title_str).font = font_sec_hdr
    ws_an.row_dimensions[start_bud_row].height = 28

    headers_b = [
        L["col_category"],
        {"ru": "Лимит", "en": "Limit", "kk": "Лимит"}.get(lang),
        L["col_spent"],
        {"ru": "Осталось", "en": "Remaining", "kk": "Қалды"}.get(lang),
        {"ru": "Использовано", "en": "Used %", "kk": "Қолданылды"}.get(lang)
    ]
    for c_idx, h in enumerate(headers_b, start=8):  # H, I, J, K, L
        cell = ws_an.cell(row=start_bud_row + 1, column=c_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_all
    ws_an.row_dimensions[start_bud_row + 1].height = 24

    b_row_idx = start_bud_row + 2
    if budgets_data:
        tot_b_limit = 0
        tot_b_spent = 0
        for brow in budgets_data:
            b_name, b_emoji, b_limit, b_spent = brow
            b_limit = int(b_limit or 0)
            b_spent = int(b_spent or 0)
            b_rem = b_limit - b_spent
            b_pct = b_spent / b_limit if b_limit > 0 else 0.0
            
            b_translated = t_category(b_name, lang)
            b_display = f"{b_emoji} {b_translated}".strip() if b_emoji or b_translated else b_name
            
            ws_an.cell(row=b_row_idx, column=8, value=b_display).alignment = Alignment(horizontal="left", indent=1)
            ws_an.cell(row=b_row_idx, column=9, value=b_limit)
            ws_an.cell(row=b_row_idx, column=10, value=b_spent)
            ws_an.cell(row=b_row_idx, column=11, value=b_rem)
            
            # Progress bar
            num_blocks = min(10, max(0, int(round(b_pct * 10))))
            bar_str = "█" * num_blocks + "░" * (10 - num_blocks)
            ws_an.cell(row=b_row_idx, column=12, value=f"{bar_str}  {b_pct * 100:.1f}%")
            
            ws_an.cell(row=b_row_idx, column=9).number_format = '#,##0'
            ws_an.cell(row=b_row_idx, column=10).number_format = '#,##0'
            ws_an.cell(row=b_row_idx, column=11).number_format = '#,##0'
            ws_an.cell(row=b_row_idx, column=12).alignment = Alignment(horizontal="left")
            
            for c in range(8, 13):
                cell = ws_an.cell(row=b_row_idx, column=c)
                cell.font = font_data
                cell.border = border_all
                if b_row_idx % 2 == 1:
                    cell.fill = fill_zebra
            ws_an.row_dimensions[b_row_idx].height = 20
            
            tot_b_limit += b_limit
            tot_b_spent += b_spent
            b_row_idx += 1
            
        tot_b_rem = tot_b_limit - tot_b_spent
        tot_b_pct = tot_b_spent / tot_b_limit if tot_b_limit > 0 else 0.0
        
        ws_an.cell(row=b_row_idx, column=8, value=L["total"]).font = font_bold
        ws_an.cell(row=b_row_idx, column=8).alignment = Alignment(horizontal="center")
        ws_an.cell(row=b_row_idx, column=9, value=tot_b_limit).font = font_bold
        ws_an.cell(row=b_row_idx, column=10, value=tot_b_spent).font = font_bold
        ws_an.cell(row=b_row_idx, column=11, value=tot_b_rem).font = font_bold
        
        num_blocks = min(10, max(0, int(round(tot_b_pct * 10))))
        tot_bar_str = "█" * num_blocks + "░" * (10 - num_blocks)
        ws_an.cell(row=b_row_idx, column=12, value=f"{tot_bar_str}  {tot_b_pct * 100:.1f}%").font = font_bold
        
        ws_an.cell(row=b_row_idx, column=9).number_format = '#,##0'
        ws_an.cell(row=b_row_idx, column=10).number_format = '#,##0'
        ws_an.cell(row=b_row_idx, column=11).number_format = '#,##0'
        
        for c in range(8, 13):
            ws_an.cell(row=b_row_idx, column=c).border = border_double_bottom
        ws_an.row_dimensions[b_row_idx].height = 22
        end_bud_row = b_row_idx
    else:
        # Placeholder row
        no_budgets_text = {
            "ru": "Лимиты бюджетов не установлены",
            "en": "No budgets limits set",
            "kk": "Бюджет лимиттері орнатылмаған"
        }.get(lang)
        ws_an.merge_cells(start_row=b_row_idx, start_column=8, end_row=b_row_idx, end_column=12)
        cell_empty = ws_an.cell(row=b_row_idx, column=8, value=no_budgets_text)
        cell_empty.font = font_subtext
        cell_empty.alignment = Alignment(horizontal="center", vertical="center")
        for c in range(8, 13):
            ws_an.cell(row=b_row_idx, column=c).border = border_all
        ws_an.row_dimensions[b_row_idx].height = 20
        end_bud_row = b_row_idx

    # 3. Month-End Prediction (Columns H:L)
    start_pred_row = end_bud_row + 2
    ws_an.cell(row=start_pred_row, column=8, value=L["prediction_title"]).font = font_sec_hdr
    ws_an.row_dimensions[start_pred_row].height = 28

    now_local = datetime.now(ZoneInfo(tz_name))
    days_in_month = calendar.monthrange(now_local.year, now_local.month)[1]
    day_of_month = max(1, now_local.day)
    curr_month_str = now_local.strftime("%Y-%m")
    curr_month_exp = monthly_data[curr_month_str]["expense"] if curr_month_str in monthly_data else total_expense
    projected_spending = int((curr_month_exp / day_of_month) * days_in_month)

    ws_an.merge_cells(start_row=start_pred_row + 1, start_column=8, end_row=start_pred_row + 1, end_column=10)
    ws_an.cell(row=start_pred_row + 1, column=8, value=L["projected_spending"]).font = font_data
    ws_an.cell(row=start_pred_row + 1, column=8).border = border_bottom
    ws_an.cell(row=start_pred_row + 1, column=9).border = border_bottom
    ws_an.cell(row=start_pred_row + 1, column=10).border = border_bottom

    cell_proj = ws_an.cell(row=start_pred_row + 1, column=12, value=projected_spending)
    cell_proj.font = font_mono
    cell_proj.alignment = Alignment(horizontal="right")
    cell_proj.number_format = f'#,##0" {currency}"'
    cell_proj.border = border_bottom
    ws_an.cell(row=start_pred_row + 1, column=11).border = border_bottom
    ws_an.cell(row=start_pred_row + 1, column=12).border = border_bottom
    ws_an.row_dimensions[start_pred_row + 1].height = 22

    # 4. AI Category Analysis (Columns H:L)
    start_ai_cat_row = start_pred_row + 3
    ws_an.cell(row=start_ai_cat_row, column=8, value=L["ai_analytic_insight"]).font = font_sec_hdr
    ws_an.row_dimensions[start_ai_cat_row].height = 28

    category_insights = []
    if all_insights:
        for ins in all_insights:
            key = ins.get("key", "").lower()
            if any(k in key for k in ["category", "spike", "budget", "expense"]):
                category_insights.append(ins.get("text"))
    
    if category_insights:
        ai_cat_text = "💡 " + "\n\n💡 ".join(category_insights)
    else:
        ai_cat_text = {
            "ru": "💡 Анализ трат по категориям не выявил критических отклонений. Все расходы находятся в пределах нормы.",
            "en": "💡 Category spending analysis did not reveal any critical deviations. All expenses are within normal ranges.",
            "kk": "💡 Санаттар бойынша шығыстарды талдау ешқандай критикалық ауытқуларды анықтаған жоқ. Барлық шығыстар қалыпты шегінде."
        }.get(lang)

    ws_an.merge_cells(start_row=start_ai_cat_row + 1, start_column=8, end_row=start_ai_cat_row + 3, end_column=12)
    cell_ai_cat = ws_an.cell(row=start_ai_cat_row + 1, column=8, value=ai_cat_text)
    cell_ai_cat.font = font_data
    cell_ai_cat.fill = fill_accent_green
    cell_ai_cat.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
    
    for r in range(start_ai_cat_row + 1, start_ai_cat_row + 4):
        for c in range(8, 13):
            ws_an.cell(row=r, column=c).border = border_all
            ws_an.cell(row=r, column=c).fill = fill_accent_green
    ws_an.row_dimensions[start_ai_cat_row + 1].height = 20
    ws_an.row_dimensions[start_ai_cat_row + 2].height = 20
    ws_an.row_dimensions[start_ai_cat_row + 3].height = 20
    
    end_right_row = start_ai_cat_row + 3

    # ==========================================
    # ANOMALIES (Merged Columns A:L)
    # ==========================================
    start_anomaly_row = max(end_left_row, end_right_row) + 3
    ws_an.cell(row=start_anomaly_row, column=1, value=L["anomalies_title"]).font = font_sec_hdr
    ws_an.row_dimensions[start_anomaly_row].height = 28

    anomalies = []
    for cat_name, spent in sorted_cats:
        if total_category_spent > 0:
            share = spent / total_category_spent
            if share > 0.45:
                anomalies.append(
                    f"Категория {cat_name} составляет {share * 100:.1f}% расходов"
                    if lang == "ru" else
                    (f"{cat_name} санатасы шығыстардың {share * 100:.1f}% құрайды" if lang == "kk" else f"{cat_name} accounts for {share * 100:.1f}% of expenses")
                )
                
    if not anomalies:
        anomalies.append(L["no_anomalies"])
        
    a_row = start_anomaly_row + 1
    for anomaly in anomalies:
        ws_an.merge_cells(start_row=a_row, start_column=1, end_row=a_row, end_column=12)
        cell_an = ws_an.cell(row=a_row, column=1, value=f"• {anomaly}")
        cell_an.font = font_subtext
        cell_an.alignment = Alignment(vertical="center", indent=1)
        ws_an.row_dimensions[a_row].height = 20
        a_row += 1

    # Widths for Analytics Sheet to prevent truncation
    widths_an = {"A": 24, "B": 15, "C": 15, "D": 15, "E": 15, "F": 4, "G": 4, "H": 25, "I": 18, "J": 18, "K": 18, "L": 24}
    for col, w in widths_an.items():
        ws_an.column_dimensions[col].width = w

    # ==========================================
    # SHEET 3: DEBTS & SCHEDULES (Credits, Recurring, Planned)
    # ==========================================
    ws_sch = wb.create_sheet(title=L["sheet_schedules"])
    ws_sch.views.sheetView[0].showGridLines = False

    # 1. Debts & Liabilities table (Columns A:F)
    ws_sch.cell(row=1, column=1, value=L["liabilities_title"]).font = font_sec_hdr
    ws_sch.row_dimensions[1].height = 30

    headers_liab = {
        "ru": ["Название обязательства", "Тип", "Остаток долга", "Сумма платежа", "Дата платежа", "Примечание"],
        "en": ["Liability Title", "Type", "Remaining Debt", "Payment Amount", "Due Date", "Note"],
        "kk": ["Міндеттеме атауы", "Түрі", "Қарыз қалдығы", "Төлем сомасы", "Төлем күні", "Түсініктеме"]
    }.get(lang)

    for c_idx, h in enumerate(headers_liab, start=1):
        cell = ws_sch.cell(row=2, column=c_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_all
    ws_sch.row_dimensions[2].height = 24

    liab_kind_map = {
        "ru": {"private_out": "Мне должны", "private_in": "Мой долг", "loan": "Кредит", "credit_card": "Кр. карта"},
        "en": {"private_out": "Owed to me", "private_in": "My debt", "loan": "Loan", "credit_card": "Credit Card"},
        "kk": {"private_out": "Маған қарыз", "private_in": "Менің қарызым", "loan": "Несие", "credit_card": "Несиелік карта"}
    }.get(lang, {"private_out": "Owed to me", "private_in": "My debt", "loan": "Loan", "credit_card": "Credit Card"})

    l_row_idx = 3
    if liabilities_data:
        for lrow in liabilities_data:
            l_kind, l_title, l_rem, l_pay, l_date, l_note = lrow
            l_rem = int(l_rem or 0)
            l_pay = int(l_pay or 0) if l_pay is not None else 0
            
            l_kind_str = liab_kind_map.get(str(l_kind or "").strip().lower(), str(l_kind or ""))
            
            ws_sch.cell(row=l_row_idx, column=1, value=l_title).alignment = Alignment(horizontal="left", indent=1)
            ws_sch.cell(row=l_row_idx, column=2, value=l_kind_str).alignment = Alignment(horizontal="center")
            ws_sch.cell(row=l_row_idx, column=3, value=l_rem)
            ws_sch.cell(row=l_row_idx, column=4, value=l_pay)
            ws_sch.cell(row=l_row_idx, column=5, value=str(l_date or "")).alignment = Alignment(horizontal="center")
            ws_sch.cell(row=l_row_idx, column=6, value=str(l_note or "")).alignment = Alignment(horizontal="left")
            
            ws_sch.cell(row=l_row_idx, column=3).number_format = '#,##0'
            ws_sch.cell(row=l_row_idx, column=4).number_format = '#,##0'
            
            for c in range(1, 7):
                cell = ws_sch.cell(row=l_row_idx, column=c)
                cell.font = font_data
                cell.border = border_all
                if l_row_idx % 2 == 1:
                    cell.fill = fill_zebra
            ws_sch.row_dimensions[l_row_idx].height = 20
            l_row_idx += 1
            
        ws_sch.cell(row=l_row_idx, column=1, value=L["total"]).font = font_bold
        ws_sch.cell(row=l_row_idx, column=1).alignment = Alignment(horizontal="center")
        
        tot_l_rem = sum(int(r[2] or 0) for r in liabilities_data)
        tot_l_pay = sum(int(r[3] or 0) for r in liabilities_data if r[3] is not None)
        
        ws_sch.cell(row=l_row_idx, column=3, value=tot_l_rem).font = font_bold
        ws_sch.cell(row=l_row_idx, column=4, value=tot_l_pay).font = font_bold
        ws_sch.cell(row=l_row_idx, column=3).number_format = '#,##0'
        ws_sch.cell(row=l_row_idx, column=4).number_format = '#,##0'
        
        for c in range(1, 7):
            ws_sch.cell(row=l_row_idx, column=c).border = border_double_bottom
        ws_sch.row_dimensions[l_row_idx].height = 22
        last_liab_row = l_row_idx
    else:
        no_liab_text = {
            "ru": "Нет активных кредитов или долгов",
            "en": "No active debts or liabilities",
            "kk": "Белсенді несиелер немесе борыштар жоқ"
        }.get(lang)
        ws_sch.merge_cells(start_row=l_row_idx, start_column=1, end_row=l_row_idx, end_column=6)
        cell_empty = ws_sch.cell(row=l_row_idx, column=1, value=no_liab_text)
        cell_empty.font = font_subtext
        cell_empty.alignment = Alignment(horizontal="center", vertical="center")
        for c in range(1, 7):
            ws_sch.cell(row=l_row_idx, column=c).border = border_all
        ws_sch.row_dimensions[l_row_idx].height = 20
        last_liab_row = l_row_idx

    # 2. Recurring payments and subscriptions (Columns H:L)
    ws_sch.cell(row=1, column=8, value=L["recurring_title"]).font = font_sec_hdr

    headers_rec = {
        "ru": ["Название шаблона", "Сумма", "Тип", "День списания", "Комментарий"],
        "en": ["Template Name", "Amount", "Type", "Day of Month", "Comment"],
        "kk": ["Үлгі атауы", "Сомасы", "Түрі", "Төлем күні", "Түсініктеме"]
    }.get(lang)

    for c_idx, h in enumerate(headers_rec, start=8):  # H, I, J, K, L
        cell = ws_sch.cell(row=2, column=c_idx, value=h)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_all
    ws_sch.row_dimensions[2].height = 24

    rec_type_map = {
        "ru": {"expense": "Расход", "income": "Доход"},
        "en": {"expense": "Expense", "income": "Income"},
        "kk": {"expense": "Шығыс", "income": "Кіріс"}
    }.get(lang, {"expense": "Expense", "income": "Income"})

    r_row_idx = 3
    if recurring_data:
        for rrow in recurring_data:
            r_title, r_amt, r_type, r_day, r_comment, r_next = rrow
            r_amt = int(r_amt or 0)
            
            r_type_str = rec_type_map.get(str(r_type or "").strip().lower(), str(r_type or ""))
            
            r_schedule = {
                "ru": f"каждый {r_day}-й день",
                "en": f"day {r_day} monthly",
                "kk": f"әр айдың {r_day}-күні"
            }.get(lang, f"day {r_day} monthly")
            
            ws_sch.cell(row=r_row_idx, column=8, value=r_title).alignment = Alignment(horizontal="left", indent=1)
            ws_sch.cell(row=r_row_idx, column=9, value=r_amt)
            ws_sch.cell(row=r_row_idx, column=10, value=r_type_str).alignment = Alignment(horizontal="center")
            ws_sch.cell(row=r_row_idx, column=11, value=r_schedule).alignment = Alignment(horizontal="center")
            ws_sch.cell(row=r_row_idx, column=12, value=str(r_comment or "")).alignment = Alignment(horizontal="left")
            
            ws_sch.cell(row=r_row_idx, column=9).number_format = '#,##0'
            
            for c in range(8, 13):
                cell = ws_sch.cell(row=r_row_idx, column=c)
                cell.font = font_data
                cell.border = border_all
                if r_row_idx % 2 == 1:
                    cell.fill = fill_zebra
            ws_sch.row_dimensions[r_row_idx].height = 20
            r_row_idx += 1
            
        ws_sch.cell(row=r_row_idx, column=8, value=L["total"]).font = font_bold
        ws_sch.cell(row=r_row_idx, column=8).alignment = Alignment(horizontal="center")
        
        tot_r_exp = sum(int(r[1] or 0) for r in recurring_data if r[2] == 'expense')
        ws_sch.cell(row=r_row_idx, column=9, value=tot_r_exp).font = font_bold
        ws_sch.cell(row=r_row_idx, column=9).number_format = '#,##0'
        
        for c in range(8, 13):
            ws_sch.cell(row=r_row_idx, column=c).border = border_double_bottom
        ws_sch.row_dimensions[r_row_idx].height = 22
        last_rec_row = r_row_idx
    else:
        no_rec_text = {
            "ru": "Нет регулярных платежей",
            "en": "No recurring payments",
            "kk": "Тұрақты төлемдер жоқ"
        }.get(lang)
        ws_sch.merge_cells(start_row=r_row_idx, start_column=8, end_row=r_row_idx, end_column=12)
        cell_empty = ws_sch.cell(row=r_row_idx, column=8, value=no_rec_text)
        cell_empty.font = font_subtext
        cell_empty.alignment = Alignment(horizontal="center", vertical="center")
        for c in range(8, 13):
            ws_sch.cell(row=r_row_idx, column=c).border = border_all
        ws_sch.row_dimensions[r_row_idx].height = 20
        last_rec_row = r_row_idx

    # 3. Planned operations (Columns A:L)
    start_planned_row = max(last_liab_row, last_rec_row) + 3
    ws_sch.cell(row=start_planned_row, column=1, value=L["planned_title"]).font = font_sec_hdr
    ws_sch.row_dimensions[start_planned_row].height = 28

    headers_plan = {
        "ru": ["Название цели", "Тип", "Сумма", "Планируемая дата", "Счёт", "Категория", "Обязательно", "Комментарий"],
        "en": ["Goal Title", "Type", "Amount", "Planned Date", "Account", "Category", "Mandatory", "Comment"],
        "kk": ["Мақсат атауы", "Түрі", "Сомасы", "Жоспарланған күн", "Шот", "Санат", "Міндетті", "Түсініктеме"]
    }.get(lang)

    ws_sch.merge_cells(start_row=start_planned_row + 1, start_column=1, end_row=start_planned_row + 1, end_column=2)
    ws_sch.cell(row=start_planned_row + 1, column=1, value=headers_plan[0])
    
    col_mappings_plan = [
        (3, headers_plan[1]),
        (4, headers_plan[2]),
        (5, headers_plan[3]),
        (6, headers_plan[4]),
        (8, headers_plan[5]),
        (9, headers_plan[6]),
        (10, headers_plan[7])
    ]
    
    for c_idx, h in col_mappings_plan:
        if c_idx == 10:
            ws_sch.merge_cells(start_row=start_planned_row + 1, start_column=10, end_row=start_planned_row + 1, end_column=12)
        elif c_idx == 6:
            ws_sch.merge_cells(start_row=start_planned_row + 1, start_column=6, end_row=start_planned_row + 1, end_column=7)
        ws_sch.cell(row=start_planned_row + 1, column=c_idx, value=h)
        
    for c in range(1, 13):
        cell = ws_sch.cell(row=start_planned_row + 1, column=c)
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_all
    ws_sch.row_dimensions[start_planned_row + 1].height = 24

    p_row_idx = start_planned_row + 2
    if planned_data:
        for prow in planned_data:
            p_title, p_kind, p_amt, p_date, p_acc, p_cat_name, p_cat_emoji, p_req, p_comment = prow
            p_amt = int(p_amt or 0)
            
            p_kind_str = rec_type_map.get(str(p_kind or "").strip().lower(), str(p_kind or ""))
            p_req_str = {
                "ru": "Да" if p_req else "Нет",
                "en": "Yes" if p_req else "No",
                "kk": "Иә" if p_req else "Жоқ"
            }.get(lang, "Yes" if p_req else "No")
            
            p_translated_cat = t_category(p_cat_name, lang)
            p_cat_display = f"{p_cat_emoji} {p_translated_cat}".strip() if p_cat_emoji or p_translated_cat else p_cat_name
            
            ws_sch.merge_cells(start_row=p_row_idx, start_column=1, end_row=p_row_idx, end_column=2)
            ws_sch.cell(row=p_row_idx, column=1, value=p_title).alignment = Alignment(horizontal="left", indent=1)
            ws_sch.cell(row=p_row_idx, column=3, value=p_kind_str).alignment = Alignment(horizontal="center")
            ws_sch.cell(row=p_row_idx, column=4, value=p_amt)
            ws_sch.cell(row=p_row_idx, column=5, value=str(p_date or "")).alignment = Alignment(horizontal="center")
            
            ws_sch.merge_cells(start_row=p_row_idx, start_column=6, end_row=p_row_idx, end_column=7)
            ws_sch.cell(row=p_row_idx, column=6, value=str(p_acc or "")).alignment = Alignment(horizontal="left", indent=1)
            
            ws_sch.cell(row=p_row_idx, column=8, value=p_cat_display).alignment = Alignment(horizontal="left", indent=1)
            ws_sch.cell(row=p_row_idx, column=9, value=p_req_str).alignment = Alignment(horizontal="center")
            
            ws_sch.merge_cells(start_row=p_row_idx, start_column=10, end_row=p_row_idx, end_column=12)
            ws_sch.cell(row=p_row_idx, column=10, value=str(p_comment or "")).alignment = Alignment(horizontal="left", indent=1)
            
            ws_sch.cell(row=p_row_idx, column=4).number_format = '#,##0'
            
            for c in range(1, 13):
                cell = ws_sch.cell(row=p_row_idx, column=c)
                cell.font = font_data
                cell.border = border_all
                if p_row_idx % 2 == 1:
                    cell.fill = fill_zebra
            ws_sch.row_dimensions[p_row_idx].height = 20
            p_row_idx += 1
            
        ws_sch.merge_cells(start_row=p_row_idx, start_column=1, end_row=p_row_idx, end_column=3)
        ws_sch.cell(row=p_row_idx, column=1, value=L["total"]).font = font_bold
        ws_sch.cell(row=p_row_idx, column=1).alignment = Alignment(horizontal="center")
        
        tot_p_amt = sum(int(r[2] or 0) for r in planned_data if r[1] == 'expense')
        ws_sch.cell(row=p_row_idx, column=4, value=tot_p_amt).font = font_bold
        ws_sch.cell(row=p_row_idx, column=4).number_format = '#,##0'
        
        for c in range(1, 13):
            cell = ws_sch.cell(row=p_row_idx, column=c)
            cell.border = border_double_bottom
            if c in [1, 2, 3]:
                cell.font = font_bold
        ws_sch.row_dimensions[p_row_idx].height = 22
    else:
        no_plan_text = {
            "ru": "Нет запланированных операций",
            "en": "No planned transactions",
            "kk": "Жоспарланған операциялар жоқ"
        }.get(lang)
        ws_sch.merge_cells(start_row=p_row_idx, start_column=1, end_row=p_row_idx, end_column=12)
        cell_empty = ws_sch.cell(row=p_row_idx, column=1, value=no_plan_text)
        cell_empty.font = font_subtext
        cell_empty.alignment = Alignment(horizontal="center", vertical="center")
        for c in range(1, 13):
            ws_sch.cell(row=p_row_idx, column=c).border = border_all
        ws_sch.row_dimensions[p_row_idx].height = 20

    widths_sch = {"A": 25, "B": 18, "C": 16, "D": 16, "E": 16, "F": 20, "G": 10, "H": 25, "I": 16, "J": 16, "K": 20, "L": 25}
    for col, w in widths_sch.items():
        ws_sch.column_dimensions[col].width = w

    # ==========================================
    # SHEET 4: TRANSACTIONS (Clean raw table)
    # ==========================================
    ws_tx = wb.create_sheet(title=L["sheet_transactions"])
    ws_tx.views.sheetView[0].showGridLines = True
    ws_tx.append(L["raw_headers"])

    for cell in ws_tx[1]:
        cell.font = font_tbl_hdr
        cell.fill = fill_tbl_hdr
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_all
    ws_tx.row_dimensions[1].height = 26

    type_local_map = {
        "ru": {"income": "Доход", "expense": "Расход", "transfer": "Перевод"},
        "en": {"income": "Income", "expense": "Expense", "transfer": "Transfer"},
        "kk": {"income": "Кіріс", "expense": "Шығыс", "transfer": "Аударма"}
    }.get(lang, {"income": "Income", "expense": "Expense", "transfer": "Transfer"})

    tx_row_idx = 2
    last_month_key = None

    for r in rows_list:
        tx_id, ts, ttype, amount, account, category, emoji, note = r
        
        try:
            dt = datetime.fromisoformat(ts)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            local_dt = dt.astimezone(ZoneInfo(tz_name))
            month_key = local_dt.strftime("%Y-%m")
            month_display = format_localized_month(month_key, lang)
        except Exception:
            month_key = str(ts)[:7] if ts and len(str(ts)) >= 7 else "Unknown"
            month_display = month_key
            
        if month_key != last_month_key:
            ws_tx.merge_cells(start_row=tx_row_idx, start_column=1, end_row=tx_row_idx, end_column=8)
            cell_sep = ws_tx.cell(row=tx_row_idx, column=1, value=f"📅 {month_display.upper()}")
            cell_sep.font = Font(name="Segoe UI", size=10, bold=True, color="475569")
            cell_sep.fill = fill_tbl_hdr
            cell_sep.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            
            for col in range(1, 9):
                c_cell = ws_tx.cell(row=tx_row_idx, column=col)
                c_cell.fill = fill_tbl_hdr
                c_cell.border = Border(top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1'))
                
            ws_tx.row_dimensions[tx_row_idx].height = 26
            tx_row_idx += 1
            last_month_key = month_key

        translated_cat = t_category(category, lang)
        cat_display = f"{emoji} {translated_cat}".strip() if emoji or translated_cat else ""
        ttype_display = type_local_map.get(str(ttype or "").strip().lower(), str(ttype or ""))
        
        try:
            dt = datetime.fromisoformat(ts)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            local_dt = dt.astimezone(ZoneInfo(tz_name))
            clean_date = format_localized_datetime(local_dt, lang)
        except Exception:
            clean_date = str(ts)[:16].replace("T", " ") if ts else ""
        
        row_vals = [
            int(tx_id),
            clean_date,
            ttype_display,
            int(amount or 0),
            currency,
            str(account or ""),
            cat_display,
            str(note or ""),
        ]

        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws_tx.cell(row=tx_row_idx, column=col_idx, value=val)
            cell.font = font_data
            cell.border = border_all
            if tx_row_idx % 2 == 1:
                cell.fill = fill_zebra
                
            if col_idx in [1, 2, 3, 5]:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 4:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0'
            else:
                cell.alignment = Alignment(horizontal="left")
            
        ws_tx.row_dimensions[tx_row_idx].height = 20
        tx_row_idx += 1

    widths_tx = [10, 26, 14, 15, 10, 18, 22, 40]
    for col_idx, width in enumerate(widths_tx, start=1):
        ws_tx.column_dimensions[chr(ord("A") + col_idx - 1)].width = width

    # Output to in-memory bytes
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_csv(rows: Iterable[tuple], lang: str, currency: str) -> bytes:
    headers_by_lang = {
        "ru": ["ID", "Дата (UTC)", "Тип", "Сумма", "Валюта", "Счёт", "Категория", "Комментарий"],
        "en": ["ID", "Date (UTC)", "Type", "Amount", "Currency", "Account", "Category", "Note"],
        "kk": ["ID", "Күні (UTC)", "Түрі", "Сома", "Валюта", "Шот", "Санат", "Түсініктеме"],
    }
    headers = headers_by_lang.get(lang, headers_by_lang["ru"])

    buf = io.StringIO()
    writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
    writer.writerow(headers)
    for r in rows:
        tx_id, ts, ttype, amount, account, category, emoji, note = r
        translated_cat = t_category(category, lang)
        cat_display = f"{emoji} {translated_cat}".strip() if emoji or translated_cat else ""
        writer.writerow([tx_id, ts, ttype, int(amount or 0), currency, account, cat_display, note])
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


async def send_premium_xlsx_report(bot, db: aiosqlite.Connection, user_id: int, period: str, lang: str, chat_id: int):
    # Fetch user settings to get currency and timezone
    settings = await get_settings(db, user_id)
    currency = _row_get(settings, "currency", 0, default="KZT")
    tz_name = _row_get(settings, "timezone", 1, default="Asia/Aqtobe")

    try:
        start_iso, end_iso, label = await _resolve_period(db, user_id, period)
    except Exception:
        from app.ui.i18n import t as _t
        await bot.send_message(chat_id, _t(lang, "EXPORT_PERIOD_ERROR"))
        return

    rows = await _fetch_rows(db, user_id, start_iso, end_iso)
    if not rows:
        empty_text = {
            "ru": "За выбранный период нет операций.",
            "en": "No transactions for the selected period.",
            "kk": "Таңдалған кезеңде операциялар жоқ.",
        }.get(lang, "За выбранный период нет операций.")
        await bot.send_message(chat_id, empty_text)
        return

    # Fetch additional metrics and database records for Executive Summary
    from app.domain.services.financial_analysis_engine import calculate_financial_metrics
    from app.domain.services.ai_priority_engine import select_top_insights
    
    metrics = await calculate_financial_metrics(db, user_id, tz_name)
    
    cur_prof = await db.execute("SELECT user_stage, behavioral_summary, discipline_score FROM ai_profile WHERE user_id=?", (user_id,))
    row_prof = await cur_prof.fetchone()
    if row_prof:
        profile = {"stage": row_prof[0], "behavioral_summary": row_prof[1], "discipline_score": row_prof[2]}
    else:
        profile = {"stage": "chaotic", "behavioral_summary": "накапливаем статистику", "discipline_score": 100}
        
    cur_ins = await db.execute("SELECT insight_key, insight_text, confidence FROM ai_insights WHERE user_id=? AND status='active'", (user_id,))
    rows_ins = await cur_ins.fetchall()
    ai_insights = [{"key": r[0], "text": r[1], "confidence": r[2]} for r in rows_ins]
    priority_insights = select_top_insights(ai_insights)
    
    cur_rec = await db.execute(
        """
        SELECT recommendation_type, message_text, target_metric_name, target_metric_start_value, target_metric_goal_value 
        FROM ai_recommendations_log 
        WHERE user_id=? AND status='sent' 
        ORDER BY id DESC LIMIT 1
        """,
        (user_id,)
    )
    row_rec = await cur_rec.fetchone()
    if row_rec:
        latest_rec = {
            "type": row_rec[0],
            "text": row_rec[1],
            "metric": row_rec[2],
            "start": row_rec[3],
            "goal": row_rec[4]
        }
    else:
        latest_rec = None

    # Fetch details for Sheet 1, Sheet 2, Sheet 3
    cur_acc = await db.execute(
        """
        SELECT id, name, balance, starting_balance, currency, is_saving
        FROM accounts
        WHERE user_id = ? AND is_archived = 0
        """,
        (user_id,)
    )
    accounts_data = await cur_acc.fetchall()

    now_local = await now_in_user_tz(db, user_id)
    curr_month_str = now_local.strftime("%Y-%m")
    
    from app.domain.services.reports_service import month_bounds_utc, iso
    start_utc, end_utc, _, _ = month_bounds_utc(tz_name, datetime.now(timezone.utc))
    start_iso_curr = iso(start_utc)
    end_iso_curr = iso(end_utc)

    cur_bud = await db.execute(
        """
        SELECT c.name, c.emoji, b.limit_amount,
               COALESCE((
                   SELECT SUM(t.amount)
                   FROM transactions t
                   WHERE t.user_id = b.user_id
                     AND t.category_id = b.category_id
                     AND t.type = 'expense'
                     AND t.deleted_at IS NULL
                     AND t.ts >= ? AND t.ts < ?
               ), 0) AS spent
        FROM budgets b
        JOIN categories c ON c.id = b.category_id
        WHERE b.user_id = ? AND b.month = ?
        """,
        (start_iso_curr, end_iso_curr, user_id, curr_month_str)
    )
    budgets_data = await cur_bud.fetchall()

    liabilities_data = []
    try:
        cur_liab = await db.execute(
            """
            SELECT kind, title, remaining_amount, payment_amount, next_payment_date, note
            FROM liabilities
            WHERE user_id = ? AND status = 'active'
            """,
            (user_id,)
        )
        liabilities_data = await cur_liab.fetchall()
    except Exception:
        try:
            cur_deb = await db.execute(
                """
                SELECT direction AS kind, title, remaining_amount, payment_amount, next_payment_date, note
                FROM debts
                WHERE user_id = ? AND is_active = 1
                """,
                (user_id,)
            )
            liabilities_data = await cur_deb.fetchall()
        except Exception:
            pass

    recurring_data = []
    try:
        cur_rec = await db.execute(
            """
            SELECT title, amount, 'expense' AS rtype, day_of_month, comment, next_run_date
            FROM recurring_expenses
            WHERE user_id = ? AND is_archived = 0
            UNION ALL
            SELECT title, amount, 'income' AS rtype, day_of_month, comment, next_run_date
            FROM recurring_incomes
            WHERE user_id = ? AND is_archived = 0
            ORDER BY day_of_month ASC
            """,
            (user_id, user_id)
        )
        recurring_data = await cur_rec.fetchall()
    except Exception:
        pass

    planned_data = []
    try:
        cur_plan = await db.execute(
            """
            SELECT p.title, p.kind, p.amount, p.planned_date, a.name AS account_name, c.name AS category_name, c.emoji AS category_emoji, p.is_required, p.comment
            FROM planned_transactions p
            LEFT JOIN accounts a ON a.id = p.account_id
            LEFT JOIN categories c ON c.id = p.category_id
            WHERE p.user_id = ? AND p.is_archived = 0
            ORDER BY date(p.planned_date) ASC
            """,
            (user_id,)
        )
        planned_data = await cur_plan.fetchall()
    except Exception:
        pass

    payload = _build_xlsx(
        rows, lang, currency, user_id, metrics, profile, priority_insights, latest_rec, tz_name,
        accounts_data=accounts_data,
        budgets_data=budgets_data,
        liabilities_data=liabilities_data,
        recurring_data=recurring_data,
        planned_data=planned_data,
        all_insights=ai_insights
    )
    
    if payload is not None:
        filename = f"finance_{label}.xlsx"
        caption = {
            "ru": (
                "📊 <b>Ваш премиум-отчет готов!</b>\n\n"
                "Файл содержит 4 аналитические вкладки (переключайтесь между ними внизу документа):\n"
                "1️⃣ <b>Главная сводка</b> — Главный AI-дашборд, балансы и оценка финансового здоровья\n"
                "2️⃣ <b>Аналитика</b> — Бюджеты, сравнение трендов, прогноз и AI-анализ категорий\n"
                "3️⃣ <b>Долги и расписание</b> — Кредиты, долги, регулярные платежи и планы\n"
                "4️⃣ <b>История операций</b> — Детальный реестр транзакций с разделителями по месяцам"
            ),
            "en": (
                "📊 <b>Your premium report is ready!</b>\n\n"
                "The file contains 4 analytical sheets (switch between them at the bottom of the document):\n"
                "1️⃣ <b>Executive Summary</b> — Main AI dashboard, balances, and Financial Health score\n"
                "2️⃣ <b>Analytics</b> — Budgets, trends comparison, prediction, and AI category analysis\n"
                "3️⃣ <b>Debts & Schedules</b> — Liabilities, recurring bills, and plans\n"
                "4️⃣ <b>Transaction History</b> — Full transaction ledger with monthly separators"
            ),
            "kk": (
                "📊 <b>Сіздің премиум есебіңіз дайын!</b>\n\n"
                "Файл 4 талдау парағынан тұрады (құжаттың төменгі жағында ауысыңыз):\n"
                "1️⃣ <b>Негізгі жиынтық</b> — Басты AI-дашборд, баланстар және қаржылық денсаулық индексі\n"
                "2️⃣ <b>Талдау</b> — Бюджеттер, трендтер, болжам және AI санаттарды талдау\n"
                "3️⃣ <b>Борыштар мен кесте</b> — Несиелер, борыштар, тұрақты төлемдер және жоспарлар\n"
                "4️⃣ <b>Операциялар тарихы</b> — Ай бойынша бөлінген барлық транзакциялар тізімі"
            )
        }.get(lang, "📊 <b>Ваш премиум-отчет готов!</b>")

        await bot.send_document(
            chat_id=chat_id,
            document=BufferedInputFile(payload, filename=filename),
            caption=caption,
            parse_mode="HTML"
        )


@router.callback_query(F.data.startswith("export:buy_single:"))
async def export_buy_single(c: CallbackQuery, state: FSMContext, db: aiosqlite.Connection):
    from app.handlers.common import neutralize_keyboard
    await neutralize_keyboard(c)
    period = c.data.split(":")[-1]
    lang = await get_lang(db, c.from_user.id)
    
    title = {
        "ru": "Разовый Excel-отчет",
        "en": "Single Excel Report",
        "kk": "Бір реттік Excel есеп"
    }.get(lang, "Разовый Excel-отчет")
    
    desc = {
        "ru": f"Подробный Excel-отчет с графиками и AI-аналитикой за выбранный период ({period})",
        "en": f"Detailed Excel report with charts and AI analytics for ({period})",
        "kk": f"Таңдалған кезең үшін графиктері мен AI-талдауы бар Excel есеп ({period})"
    }.get(lang, "Подробный Excel-отчет")
    
    await c.bot.send_invoice(
        chat_id=c.from_user.id,
        title=title,
        description=desc[:255],
        payload=f"export_single:{period}",
        provider_token="",
        currency="XTR",
        prices=[
            LabeledPrice(
                label=title,
                amount=20,
            )
        ],
    )
    await c.answer()


def _export_menu_kb(lang: str, cancel_cb: str = "cancel"):
    kb = InlineKeyboardBuilder()
    labels = {
        "ru": {"day": "📅 За сегодня", "week": "📆 За неделю", "month": "🗓 За месяц", "all": "🗂 Всё", "cancel": "❌ Отмена"},
        "en": {"day": "📅 Today", "week": "📆 Week", "month": "🗓 Month", "all": "🗂 All", "cancel": "❌ Cancel"},
        "kk": {"day": "📅 Бүгін", "week": "📆 Апта", "month": "🗓 Ай", "all": "🗂 Барлығы", "cancel": "❌ Бас тарту"},
    }
    L = labels.get(lang, labels["ru"])
    for key in ("day", "week", "month", "all"):
        kb.button(text=L[key], callback_data=f"export:{key}")
    kb.button(text=L["cancel"], callback_data=cancel_cb)
    kb.adjust(2, 2, 1)
    return kb.as_markup()


@router.callback_query(F.data == "st:tx:export")
async def export_from_settings(c: CallbackQuery, state: FSMContext, db: aiosqlite.Connection):
    lang = await get_lang(db, c.from_user.id)
    prompt = {
        "ru": "📤 <b>Экспорт операций</b>\n\nВыбери период:",
        "en": "📤 <b>Export transactions</b>\n\nPick a period:",
        "kk": "📤 <b>Операцияларды экспорттау</b>\n\nКезеңді таңда:",
    }.get(lang, "📤 <b>Экспорт операций</b>\n\nВыбери период:")
    await c.message.edit_text(prompt, reply_markup=_export_menu_kb(lang, cancel_cb="st:tx_manage"), parse_mode="HTML")
    await c.answer()


async def _resolve_period(db: aiosqlite.Connection, user_id: int, period: str) -> tuple[str | None, str | None, str]:
    """Return (start_utc_iso, end_utc_iso, label) for the requested period."""
    if period == "all":
        return None, None, "all-time"

    now_local = await now_in_user_tz(db, user_id)
    settings = await get_settings(db, user_id)
    tz_name = _row_get(settings, "timezone", 1, default="Asia/Aqtobe")
    now_utc = datetime.now(timezone.utc)

    if period == "day":
        start, end, _, _ = day_bounds_utc(tz_name, now_utc)
        label = now_local.strftime("%Y-%m-%d")
    elif period == "week":
        start_day = (now_local - timedelta(days=6)).replace(hour=0, minute=0, second=0, microsecond=0)
        start = start_day.astimezone(timezone.utc)
        end = (now_local + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0).astimezone(timezone.utc)
        label = f"{start_day.strftime('%Y-%m-%d')}_{now_local.strftime('%Y-%m-%d')}"
    else:  # month
        start, end, _, _ = month_bounds_utc(tz_name, now_utc)
        label = now_local.strftime("%Y-%m")

    return iso(start), iso(end), label


@router.message(Command("export"))
async def export_entry(m: Message, db: aiosqlite.Connection):
    lang = await get_lang(db, m.from_user.id)
    prompt = {
        "ru": "📤 <b>Экспорт операций</b>\n\nВыбери период:",
        "en": "📤 <b>Export transactions</b>\n\nPick a period:",
        "kk": "📤 <b>Операцияларды экспорттау</b>\n\nКезеңді таңда:",
    }.get(lang, "📤 <b>Экспорт операций</b>\n\nВыбери период:")
    await m.answer(prompt, reply_markup=_export_menu_kb(lang), parse_mode="HTML")


@router.callback_query(F.data.startswith("export:"))
async def export_pick(c: CallbackQuery, state: FSMContext, db: aiosqlite.Connection, full_access: bool = False):
    from app.handlers.common import neutralize_keyboard
    await neutralize_keyboard(c)
    period = (c.data or "").split(":")[1]
    if period not in {"day", "week", "month", "all"}:
        await c.answer()
        return

    user_id = c.from_user.id
    lang = await get_lang(db, user_id)
    settings = await get_settings(db, user_id)
    currency = _row_get(settings, "currency", 0, default="KZT")

    try:
        start_iso, end_iso, label = await _resolve_period(db, user_id, period)
    except Exception:
        from app.ui.i18n import t as _t
        await c.answer(_t(lang, "EXPORT_PERIOD_ERROR"), show_alert=True)
        return

    rows = await _fetch_rows(db, user_id, start_iso, end_iso)
    if not rows:
        empty_text = {
            "ru": "За выбранный период нет операций.",
            "en": "No transactions for the selected period.",
            "kk": "Таңдалған кезеңде операциялар жоқ.",
        }.get(lang, "За выбранный период нет операций.")
        await c.answer(empty_text, show_alert=True)
        return

    # Acknowledge immediately so the spinner clears while we build the file.
    try:
        await c.answer({"ru": "Готовлю файл…", "en": "Building…", "kk": "Файл әзірленуде…"}.get(lang, "Готовлю файл…"))
    except Exception:
        pass

    # Determine free trial status using migration-backed users column
    exports_used = await get_free_exports_used(db, user_id)

    use_premium_xlsx = False
    is_free_trial_now = False

    if full_access:
        use_premium_xlsx = True
    elif exports_used == 0:
        # Give them exactly 1 free premium export as a trial!
        use_premium_xlsx = True
        is_free_trial_now = True

    if use_premium_xlsx:
        await send_premium_xlsx_report(c.bot, db, user_id, period, lang, c.message.chat.id)
        if is_free_trial_now:
            # Mark as used
            await increment_free_export(db, user_id)
            await db.commit()
            
            # Send celebratory 1-time free trial congratulation
            trial_msg = {
                "ru": (
                    "🎁 <b>Бесплатный Excel-отчёт с графиками</b>\n\n"
                    "Файл уже отправлен — внутри диаграммы и сводка за период.\n\n"
                    "<i>Следующие Excel-экспорты — с полным доступом в настройках.</i>"
                ),
                "en": (
                    "🎁 <b>Free Excel report with charts</b>\n\n"
                    "The file is already sent — it includes charts and a period summary.\n\n"
                    "<i>More Excel exports unlock with Full Access in settings.</i>"
                ),
                "kk": (
                    "🎁 <b>Графиктері бар тегін Excel есеп</b>\n\n"
                    "Файл жіберілді — ішінде диаграммалар мен кезең қорытындысы бар.\n\n"
                    "<i>Келесі Excel экспорттары — баптаулардағы толық қолжетімділікпен.</i>"
                ),
            }.get(lang, "🎁 <b>Бесплатный Excel-отчёт</b>")
            
            await c.message.answer(trial_msg, parse_mode="HTML")
        return

    # Fallback to CSV for free users (who already used their trial) or if openpyxl failed
    payload_csv = _build_csv(rows, lang, currency)
    filename = f"finance_{label}.csv"
    await c.message.answer_document(BufferedInputFile(payload_csv, filename=filename))
    
    # Generate the stunning preview dynamically in a separate thread (asyncio.to_thread) to avoid blocking event loop
    png_bytes = await asyncio.to_thread(_build_downgrade_preview_png, rows, lang, currency)

    buy_single_text = {
        "ru": "📊 Купить этот отчет за 20 ⭐️",
        "en": "📊 Buy this report for 20 ⭐️",
        "kk": "📊 Бұл есепті 20 ⭐️-ға сатып алу",
    }.get(lang, "📊 Купить этот отчет за 20 ⭐️")

    upgrade_btn_text = {
        "ru": "👑 Полный доступ на 3 мес за 150 ⭐️",
        "en": "👑 Full Access (3 months) for 150 ⭐️",
        "kk": "👑 Толық қолжетімділік (3 ай) - 150 ⭐️",
    }.get(lang, "👑 Полный доступ на 3 мес за 150 ⭐️")

    menu_btn_text = {
        "ru": "🏠 В Главное меню",
        "en": "🏠 Main Menu",
        "kk": "🏠 Басты мәзірге",
    }.get(lang, "🏠 В Главное меню")

    paywall_kb = InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text=buy_single_text, callback_data=f"export:buy_single:{period}")],
        [InlineKeyboardButton(text=upgrade_btn_text, callback_data="upgrade:activate")],
        [InlineKeyboardButton(text=menu_btn_text, callback_data="hub:main")]
    ])

    caption_text = {
        "ru": (
            "⚠️ <b>Бесплатный Excel-отчёт уже использован</b>\n\n"
            "На графике — ваши доходы и расходы за период. В полном доступе такие диаграммы "
            "есть в каждом Excel-экспорте.\n\n"
            "🔥 Вы можете купить этот подробный Excel-отчет разово за <b>20 ⭐️</b> или подключить полный доступ за <b>150 ⭐️</b> в настройках."
        ),
        "en": (
            "⚠️ <b>Your free Excel trial was already used</b>\n\n"
            "The chart shows your income and expenses for the period. With Full Access, "
            "charts are included in every Excel export.\n\n"
            "🔥 You can purchase this detailed Excel report as a one-time order for <b>20 ⭐️</b> or get Full Access for <b>150 ⭐️</b> in settings."
        ),
        "kk": (
            "⚠️ <b>Тегін Excel есеп қолданылған</b>\n\n"
            "Графикте кезеңдегі кіріс пен шығыстарыңыз көрсетілген. Толық қолжетімділікте "
            "әр Excel экспортында диаграммалар болады.\n\n"
            "🔥 Сіз бұл есепті бір рет <b>20 ⭐️</b>-ға сатып ала аласыз немесе баптауларда <b>150 ⭐️</b>-ға толық қолжетімділікті қоса аласыз."
        ),
    }.get(lang, "")

    if png_bytes:
        photo = BufferedInputFile(png_bytes, filename="premium_preview.png")
        await c.message.answer_photo(
            photo=photo,
            caption=caption_text,
            parse_mode="HTML",
            reply_markup=paywall_kb
        )
    else:
        await c.message.answer(
            caption_text,
            parse_mode="HTML",
            reply_markup=paywall_kb
        )
